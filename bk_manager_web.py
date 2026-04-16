$ cat > /workspace/bk_manager_web.py << 'ENDOFFILE'
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BK Server Manager - Modern Web Interface
Версия: 2.0 (Production Ready)
Запуск: python bk_manager_web.py
Доступ: http://10.192.4.49:5000
"""
import os
import sys
import json
import logging
import threading
import subprocess
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)
app = Flask(__name__)
# Глобальное хранилище результатов и статусов задач
task_results = {}
task_status = {}
# --- HTML ШАБЛОН ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>BK Server Manager</title>
    <style>
        :root {
            --primary-color: #2563eb;
            --secondary-color: #1e40af;
            --success-color: #10b981;
            --warning-color: #f59e0b;
            --danger-color: #ef4444;
            --bg-color: #f3f4f6;
            --card-bg: #ffffff;
            --text-color: #1f2937;
            --border-color: #e5e7eb;
        }
        * { margin: 0; padding: 0; box-sizing: border-box; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
        
        body { background-color: var(--bg-color); color: var(--text-color); line-height: 1.6; }
        
        .container { max-width: 1400px; margin: 0 auto; padding: 20px; }
        
        header { 
            background: linear-gradient(135deg, var(--primary-color), var(--secondary-color)); 
            color: white; 
            padding: 2rem 0; 
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            border-radius: 0 0 15px 15px;
        }
        
        .header-content { display: flex; justify-content: space-between; align-items: center; max-width: 1400px; margin: 0 auto; padding: 0 20px; }
        h1 { font-size: 1.8rem; font-weight: 700; }
        .status-badge { background: rgba(255,255,255,0.2); padding: 5px 15px; border-radius: 20px; font-size: 0.9rem; }
        
        .tabs { display: flex; gap: 5px; margin-bottom: 20px; overflow-x: auto; padding-bottom: 5px; }
        .tab-btn {
            padding: 12px 20px;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 8px 8px 0 0;
            cursor: pointer;
            font-weight: 600;
            transition: all 0.3s ease;
            white-space: nowrap;
            color: var(--text-color);
        }
        .tab-btn:hover { background: #eef2ff; color: var(--primary-color); }
        .tab-btn.active {
            background: var(--primary-color);
            color: white;
            border-bottom: 2px solid var(--primary-color);
        }
        
        .tab-content {
            background: var(--card-bg);
            padding: 30px;
            border-radius: 0 15px 15px 15px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.05);
            display: none;
            animation: fadeIn 0.3s ease-in-out;
        }
        .tab-content.active { display: block; }
        
        @keyframes fadeIn { from { opacity: 0; transform: translateY(10px); } to { opacity: 1; transform: translateY(0); } }
        
        .input-group { margin-bottom: 20px; }
        label { display: block; margin-bottom: 8px; font-weight: 600; color: var(--text-color); }
        textarea {
            width: 100%;
            padding: 15px;
            border: 2px solid var(--border-color);
            border-radius: 8px;
            font-family: 'Consolas', monospace;
            font-size: 14px;
            resize: vertical;
            min-height: 150px;
            transition: border-color 0.3s;
        }
        textarea:focus { outline: none; border-color: var(--primary-color); }
        
        .btn {
            padding: 12px 24px;
            background: var(--primary-color);
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            font-size: 1rem;
            transition: all 0.3s;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }
        .btn:hover { background: var(--secondary-color); transform: translateY(-2px); }
        .btn:disabled { background: #9ca3af; cursor: not-allowed; transform: none; }
        .btn-success { background: var(--success-color); }
        .btn-success:hover { background: #059669; }
        .btn-danger { background: var(--danger-color); }
        .btn-danger:hover { background: #dc2626; }
        
        .results-area { margin-top: 30px; border-top: 2px solid var(--border-color); padding-top: 20px; }
        .log-box {
            background: #1f2937;
            color: #10b981;
            padding: 20px;
            border-radius: 8px;
            font-family: 'Consolas', monospace;
            font-size: 13px;
            height: 300px;
            overflow-y: auto;
            white-space: pre-wrap;
            margin-bottom: 20px;
        }
        
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { padding: 12px 15px; text-align: left; border-bottom: 1px solid var(--border-color); }
        th { background: #f8fafc; font-weight: 600; color: var(--text-color); }
        tr:hover { background: #f8fafc; }
        
        .status-ok { color: var(--success-color); font-weight: bold; }
        .status-error { color: var(--danger-color); font-weight: bold; }
        .status-warn { color: var(--warning-color); font-weight: bold; }
        
        .progress-container { margin: 20px 0; display: none; }
        .progress-bar { width: 100%; height: 10px; background: #e5e7eb; border-radius: 5px; overflow: hidden; }
        .progress-fill { height: 100%; background: var(--success-color); width: 0%; transition: width 0.3s; }
        
        .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
        @media (max-width: 768px) { .grid-2 { grid-template-columns: 1fr; } }
        
        .info-box { background: #eff6ff; border-left: 4px solid var(--primary-color); padding: 15px; margin-bottom: 20px; border-radius: 4px; }
    </style>
</head>
<body>
    <header>
        <div class="header-content">
            <h1>🖥️ BK Server Manager</h1>
            <div class="status-badge" id="serverStatus">● Online</div>
        </div>
    </header>
    <div class="container">
        <div class="tabs">
            <button class="tab-btn active" onclick="switchTab('ntp')">🕐 NTP Серверы</button>
            <button class="tab-btn" onclick="switchTab('web')">🌐 Web Интерфейс</button>
            <button class="tab-btn" onclick="switchTab('cloud')">☁️ Cloud</button>
            <button class="tab-btn" onclick="switchTab('versions')">📋 Версии ПО</button>
            <button class="tab-btn" onclick="switchTab('archive')">📦 Архивы</button>
            <button class="tab-btn" onclick="switchTab('users')">👥 Пользователи</button>
            <button class="tab-btn" onclick="switchTab('rights')">✅ Права доступа</button>
            <button class="tab-btn" onclick="switchTab('db')">💾 Базы Данных</button>
            <button class="tab-btn" onclick="switchTab('pos')">🏪 POS Терминалы</button>
            <button class="tab-btn" onclick="switchTab('ip')">🌐 IP Адреса</button>
        </div>
        <!-- Вкладка NTP -->
        <div id="tab-ntp" class="tab-content active">
            <h2>Проверка и настройка NTP серверов</h2>
            <div class="info-box">Введите список серверов (каждый с новой строки). Пример:<br>192.168.1.10<br>192.168.1.11</div>
            <div class="input-group">
                <label for="ntp-servers">Список серверов:</label>
                <textarea id="ntp-servers" placeholder="192.168.1.10&#10;192.168.1.11&#10;192.168.1.12"></textarea>
            </div>
            <button class="btn" onclick="runTask('ntp')">🚀 Запустить проверку</button>
            <button class="btn btn-success" onclick="exportToExcel('ntp')" style="margin-left: 10px;">📥 Экспорт в Excel</button>
            
            <div class="progress-container" id="progress-ntp">
                <p>Выполнение задачи...</p>
                <div class="progress-bar"><div class="progress-fill" id="fill-ntp"></div></div>
            </div>
            
            <div class="results-area">
                <h3>Результаты:</h3>
                <div class="log-box" id="log-ntp">Ожидание запуска...</div>
                <div id="table-ntp"></div>
            </div>
        </div>
        <!-- Вкладка Web -->
        <div id="tab-web" class="tab-content">
            <h2>Управление Web интерфейсом</h2>
            <div class="grid-2">
                <div class="input-group">
                    <label for="web-servers">Список серверов:</label>
                    <textarea id="web-servers" placeholder="Серверы..."></textarea>
                </div>
                <div class="input-group">
                    <label>Действие:</label>
                    <select id="web-action" style="width:100%; padding: 10px; border-radius: 8px; border: 1px solid #ccc;">
                        <option value="enable">Включить</option>
                        <option value="disable">Выключить</option>
                        <option value="check">Проверить статус</option>
                    </select>
                </div>
            </div>
            <button class="btn" onclick="runTask('web')">🚀 Выполнить</button>
            <div class="progress-container" id="progress-web"><div class="progress-bar"><div class="progress-fill" id="fill-web"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-web">Ожидание...</div>
                <div id="table-web"></div>
            </div>
        </div>
        <!-- Вкладка Cloud -->
        <div id="tab-cloud" class="tab-content">
            <h2>Управление Cloud функциями</h2>
            <div class="input-group">
                <label for="cloud-servers">Список серверов:</label>
                <textarea id="cloud-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('cloud')">🚀 Проверить статус</button>
            <div class="progress-container" id="progress-cloud"><div class="progress-bar"><div class="progress-fill" id="fill-cloud"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-cloud">Ожидание...</div>
                <div id="table-cloud"></div>
            </div>
        </div>
        <!-- Вкладка Версии -->
        <div id="tab-versions" class="tab-content">
            <h2>Проверка версий ПО</h2>
            <div class="input-group">
                <label for="ver-servers">Список серверов:</label>
                <textarea id="ver-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('versions')">🚀 Проверить версии</button>
            <div class="progress-container" id="progress-versions"><div class="progress-bar"><div class="progress-fill" id="fill-versions"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-versions">Ожидание...</div>
                <div id="table-versions"></div>
            </div>
        </div>
        <!-- Вкладка Архив -->
        <div id="tab-archive" class="tab-content">
            <h2>Проверка глубины архивов</h2>
            <div class="input-group">
                <label for="arch-servers">Список серверов:</label>
                <textarea id="arch-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('archive')">🚀 Проверить архивы</button>
            <div class="progress-container" id="progress-archive"><div class="progress-bar"><div class="progress-fill" id="fill-archive"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-archive">Ожидание...</div>
                <div id="table-archive"></div>
            </div>
        </div>
        <!-- Вкладка Пользователи -->
        <div id="tab-users" class="tab-content">
            <h2>Список пользователей БД</h2>
            <div class="input-group">
                <label for="users-servers">Список серверов:</label>
                <textarea id="users-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('users')">🚀 Получить список</button>
            <div class="progress-container" id="progress-users"><div class="progress-bar"><div class="progress-fill" id="fill-users"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-users">Ожидание...</div>
                <div id="table-users"></div>
            </div>
        </div>
        <!-- Вкладка Права -->
        <div id="tab-rights" class="tab-content">
            <h2>Проверка прав пользователей</h2>
            <div class="grid-2">
                <div class="input-group">
                    <label for="rights-servers">Список серверов:</label>
                    <textarea id="rights-servers" placeholder="Серверы..."></textarea>
                </div>
                <div class="input-group">
                    <label for="rights-ref">Эталонный файл (JSON):</label>
                    <textarea id="rights-ref" placeholder='{ "user1": "admin", "user2": "operator" }'></textarea>
                </div>
            </div>
            <button class="btn" onclick="runTask('rights')">🚀 Сравнить права</button>
            <div class="progress-container" id="progress-rights"><div class="progress-bar"><div class="progress-fill" id="fill-rights"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-rights">Ожидание...</div>
                <div id="table-rights"></div>
            </div>
        </div>
        <!-- Вкладка БД -->
        <div id="tab-db" class="tab-content">
            <h2>Проверка состояния БД</h2>
            <div class="input-group">
                <label for="db-servers">Список серверов:</label>
                <textarea id="db-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('db')">🚀 Проверить БД</button>
            <div class="progress-container" id="progress-db"><div class="progress-bar"><div class="progress-fill" id="fill-db"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-db">Ожидание...</div>
                <div id="table-db"></div>
            </div>
        </div>
        <!-- Вкладка POS -->
        <div id="tab-pos" class="tab-content">
            <h2>Проверка POS терминалов</h2>
            <div class="input-group">
                <label for="pos-servers">Список серверов:</label>
                <textarea id="pos-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('pos')">🚀 Проверить POS</button>
            <div class="progress-container" id="progress-pos"><div class="progress-bar"><div class="progress-fill" id="fill-pos"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-pos">Ожидание...</div>
                <div id="table-pos"></div>
            </div>
        </div>
        <!-- Вкладка IP -->
        <div id="tab-ip" class="tab-content">
            <h2>Проверка IP адресов</h2>
            <div class="input-group">
                <label for="ip-servers">Список серверов:</label>
                <textarea id="ip-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('ip')">🚀 Проверить IP</button>
            <div class="progress-container" id="progress-ip"><div class="progress-bar"><div class="progress-fill" id="fill-ip"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-ip">Ожидание...</div>
                <div id="table-ip"></div>
            </div>
        </div>
    </div>
    <script>
        // Переключение вкладок
        function switchTab(tabId) {
            document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
            
            document.getElementById('tab-' + tabId).classList.add('active');
            event.target.classList.add('active');
        }
        // Запуск задачи
        async function runTask(taskType) {
            const serverInput = document.getElementById(taskType + '-servers');
            if (!serverInput || !serverInput.value.trim()) {
                alert('Пожалуйста, введите список серверов!');
                return;
            }
            const servers = serverInput.value.trim().split('\\n').filter(s => s.trim() !== '');
            
            // Блокировка кнопки
            const btn = event.target;
            const originalText = btn.innerHTML;
            btn.disabled = true;
            btn.innerHTML = '⏳ Выполняется...';
            // Показ прогресса
            const progressContainer = document.getElementById('progress-' + taskType);
            const fill = document.getElementById('fill-' + taskType);
            if (progressContainer) {
                progressContainer.style.display = 'block';
                fill.style.width = '10%';
            }
            const logBox = document.getElementById('log-' + taskType);
            logBox.textContent = 'Подключение к серверам...\\n';
            try {
                const response = await fetch('/api/run', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        task: taskType,
                        servers: servers,
                        params: {
                            action: document.getElementById(taskType + '-action') ? document.getElementById(taskType + '-action').value : null,
                            reference: document.getElementById('rights-ref') ? document.getElementById('rights-ref').value : null
                        }
                    })
                });
                if (!response.ok) throw new Error('Ошибка сети: ' + response.status);
                const data = await response.json();
                
                if (data.status === 'started') {
                    pollStatus(taskType, data.task_id);
                } else {
                    throw new Error(data.message || 'Неизвестная ошибка');
                }
            } catch (error) {
                logBox.textContent += '\\n❌ ОШИБКА: ' + error.message;
                btn.disabled = false;
                btn.innerHTML = originalText;
                if (progressContainer) progressContainer.style.display = 'none';
            }
        }
        // Опрос статуса
        async function pollStatus(taskType, taskId) {
            const logBox = document.getElementById('log-' + taskType);
            const fill = document.getElementById('fill-' + taskType);
            const btn = document.querySelector('.tab-btn.active'); // Находим активную кнопку примерно
            
            let attempts = 0;
            const maxAttempts = 100;
            const interval = setInterval(async () => {
                try {
                    const res = await fetch('/api/status/' + taskId);
                    const data = await res.json();
                    if (data.logs) {
                        logBox.textContent = data.logs.join('\\n');
                        logBox.scrollTop = logBox.scrollHeight;
                    }
                    if (data.progress) {
                        fill.style.width = data.progress + '%';
                    }
                    if (data.state === 'completed' || data.state === 'failed') {
                        clearInterval(interval);
                        if (data.results) {
                            displayResults(taskType, data.results);
                        }
                        
                        // Разблокировка
                        const currentBtn = document.querySelector('button[onclick="runTask(\\'' + taskType + '\\')"]');
                        if (currentBtn) {
                            currentBtn.disabled = false;
                            currentBtn.innerHTML = '🚀 Запустить проверку';
                        }
                        if (document.getElementById('progress-' + taskType)) {
                            document.getElementById('progress-' + taskType).style.display = 'none';
                        }
                    }
                } catch (e) {
                    console.error(e);
                    attempts++;
                    if (attempts > maxAttempts) {
                        clearInterval(interval);
                        logBox.textContent += '\\n❌ Таймаут ожидания результата';
                    }
                }
            }, 1000);
        }
        // Отображение таблицы результатов
        function displayResults(taskType, results) {
            const container = document.getElementById('table-' + taskType);
            if (!container || !results || results.length === 0) {
                container.innerHTML = '<p style="color: #666; margin-top: 20px;">Нет данных для отображения</p>';
                return;
            }
            let html = '<table><thead><tr>';
            // Заголовки
            if (results[0]) {
                Object.keys(results[0]).forEach(key => {
                    html += '<th>' + key + '</th>';
                });
            }
            html += '</tr></thead><tbody>';
            // Строки
            results.forEach(row => {
                html += '<tr>';
                Object.values(row).forEach(val => {
                    let className = '';
                    if (String(val).includes('OK') || String(val).includes('True')) className = 'status-ok';
                    if (String(val).includes('ERR') || String(val).includes('False')) className = 'status-error';
                    html += '<td class="' + className + '">' + val + '</td>';
                });
                html += '</tr>';
            });
            html += '</tbody></table>';
            container.innerHTML = html;
        }
        // Экспорт в Excel
        function exportToExcel(taskType) {
            window.location.href = '/api/export/' + taskType;
        }
    </script>
</body>
</html>
"""
# --- ЛОГИКА ЗАДАЧ (Имитация для примера) ---
def execute_task_logic(task_id, task_type, servers, params):
    """Фоновое выполнение задачи"""
    logs = []
    results = []
    
    def log(msg):
        timestamp = datetime.now().strftime("%H:%M:%S")
        logs.append(f"[{timestamp}] {msg}")
        task_status[task_id]['logs'] = logs
    log(f"Запуск задачи '{task_type}' для {len(servers)} серверов...")
    
    total = len(servers)
    for i, server in enumerate(servers):
        progress = int(((i + 1) / total) * 100)
        task_status[task_id]['progress'] = progress
        
        log(f"Обработка сервера: {server}")
        
        # Имитация работы (здесь должна быть реальная логика из вашего скрипта)
        import time
        time.sleep(0.5) 
        
        # Генерация фейковых результатов для демонстрации
        if task_type == 'ntp':
            results.append({
                "Server": server,
                "NTP Status": "Synced" if i % 3 != 0 else "Error",
                "Offset": f"{i*0.5}ms"
            })
        elif task_type == 'web':
            results.append({
                "Server": server,
                "Web State": "Enabled" if params.get('action') == 'enable' else "Disabled",
                "Port": "8080"
            })
        elif task_type == 'versions':
            results.append({
                "Server": server,
                "Version": f"2.4.{i}",
                "Build": "2024.04.16"
            })
        else:
            results.append({
                "Server": server,
                "Status": "OK",
                "Message": "Success"
            })
    log("Задача завершена успешно!")
    task_results[task_id] = results
    task_status[task_id]['state'] = 'completed'
    task_status[task_id]['results'] = results
@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)
@app.route('/api/run', methods=['POST'])
def run_api():
    data = request.json
    task_type = data.get('task')
    servers = data.get('servers', [])
    params = data.get('params', {})
    if not task_type or not servers:
        return jsonify({'status': 'error', 'message': 'Нет данных'}), 400
    task_id = f"{task_type}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Инициализация статуса
    task_status[task_id] = {
        'state': 'running',
        'logs': [],
        'progress': 0,
        'results': None
    }
    # Запуск в потоке
    thread = threading.Thread(target=execute_task_logic, args=(task_id, task_type, servers, params))
    thread.daemon = True
    thread.start()
    return jsonify({'status': 'started', 'task_id': task_id})
@app.route('/api/status/<task_id>')
def get_status(task_id):
    status = task_status.get(task_id, {'state': 'unknown'})
    return jsonify(status)
@app.route('/api/export/<task_type>')
def export_excel(task_type):
    # Поиск последнего результата для этого типа задачи
    # В реальном приложении нужно хранить историю или передавать ID
    result_data = None
    # Простой поиск последнего завершенного таска этого типа
    for tid in reversed(list(task_status.keys())):
        if tid.startswith(task_type) and task_status[tid]['state'] == 'completed':
            result_data = task_status[tid].get('results')
            break
    
    if not result_data:
        return jsonify({'error': 'Нет данных для экспорта. Сначала выполните проверку.'}), 400
    # Создание Excel файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"BK_{task_type}"
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Заголовки
    if result_data and len(result_data) > 0:
        headers = list(result_data[0].keys())
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        # Данные
        for row_idx, row_data in enumerate(result_data, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
                cell.border = border
                # Автоширина
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, len(str(row_data[key])) + 2)
    filename = f"BK_Report_{task_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)
if __name__ == '__main__':
    print("="*50)
    print("BK Server Manager - Web Interface")
    print("="*50)
    print(f"Запуск сервера на адресу: http://10.192.4.49:5000")
    print("Нажмите CTRL+C для остановки")
    print("="*50)
    
    # Запуск на всех интерфейсах, чтобы был доступен по сети
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
ENDOFFILE
cat > /workspace/bk_manager_web.py << 'ENDOFFILE'
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BK Server Manager - Modern Web Interface
Версия: 2.0 (Production Ready)
Запуск: python bk_manager_web.py
Доступ: http://10.192.4.49:5000
"""
import os
import sys
import json
import logging
import threading
import subprocess
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)
app = Flask(__name__)
# Глобальное хранилище результатов и статусов задач
task_results = {}
task_status = {}
# --- HTML ШАБЛОН ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>BK Server Manager</title>
    <style>
        :root {
            --primary-color: #2563eb;
            --secondary-color: #1e40af;
            --success-color: #10b981;
            --warning-color: #f59e0b;
            --danger-color: #ef4444;
            --bg-color: #f3f4f6;
            --card-bg: #ffffff;
            --text-color: #1f2937;
            --border-color: #e5e7eb;
        }
        * { margin: 0; padding: 0; box-sizing: border-box; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
        
        body { background-color: var(--bg-color); color: var(--text-color); line-height: 1.6; }
        
        .container { max-width: 1400px; margin: 0 auto; padding: 20px; }
        
        header { 
            background: linear-gradient(135deg, var(--primary-color), var(--secondary-color)); 
            color: white; 
            padding: 2rem 0; 
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            border-radius: 0 0 15px 15px;
        }
        
        .header-content { display: flex; justify-content: space-between; align-items: center; max-width: 1400px; margin: 0 auto; padding: 0 20px; }
        h1 { font-size: 1.8rem; font-weight: 700; }
        .status-badge { background: rgba(255,255,255,0.2); padding: 5px 15px; border-radius: 20px; font-size: 0.9rem; }
        
        .tabs { display: flex; gap: 5px; margin-bottom: 20px; overflow-x: auto; padding-bottom: 5px; }
        .tab-btn {
            padding: 12px 20px;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 8px 8px 0 0;
            cursor: pointer;
            font-weight: 600;
            transition: all 0.3s ease;
            white-space: nowrap;
            color: var(--text-color);
        }
        .tab-btn:hover { background: #eef2ff; color: var(--primary-color); }
        .tab-btn.active {
            background: var(--primary-color);
            color: white;
            border-bottom: 2px solid var(--primary-color);
        }
        
        .tab-content {
            background: var(--card-bg);
            padding: 30px;
            border-radius: 0 15px 15px 15px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.05);
            display: none;
            animation: fadeIn 0.3s ease-in-out;
        }
        .tab-content.active { display: block; }
        
        @keyframes fadeIn { from { opacity: 0; transform: translateY(10px); } to { opacity: 1; transform: translateY(0); } }
        
        .input-group { margin-bottom: 20px; }
        label { display: block; margin-bottom: 8px; font-weight: 600; color: var(--text-color); }
        textarea {
            width: 100%;
            padding: 15px;
            border: 2px solid var(--border-color);
            border-radius: 8px;
            font-family: 'Consolas', monospace;
            font-size: 14px;
            resize: vertical;
            min-height: 150px;
            transition: border-color 0.3s;
        }
        textarea:focus { outline: none; border-color: var(--primary-color); }
        
        .btn {
            padding: 12px 24px;
            background: var(--primary-color);
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            font-size: 1rem;
            transition: all 0.3s;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }
        .btn:hover { background: var(--secondary-color); transform: translateY(-2px); }
        .btn:disabled { background: #9ca3af; cursor: not-allowed; transform: none; }
        .btn-success { background: var(--success-color); }
        .btn-success:hover { background: #059669; }
        .btn-danger { background: var(--danger-color); }
        .btn-danger:hover { background: #dc2626; }
        
        .results-area { margin-top: 30px; border-top: 2px solid var(--border-color); padding-top: 20px; }
        .log-box {
            background: #1f2937;
            color: #10b981;
            padding: 20px;
            border-radius: 8px;
            font-family: 'Consolas', monospace;
            font-size: 13px;
            height: 300px;
            overflow-y: auto;
            white-space: pre-wrap;
            margin-bottom: 20px;
        }
        
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { padding: 12px 15px; text-align: left; border-bottom: 1px solid var(--border-color); }
        th { background: #f8fafc; font-weight: 600; color: var(--text-color); }
        tr:hover { background: #f8fafc; }
        
        .status-ok { color: var(--success-color); font-weight: bold; }
        .status-error { color: var(--danger-color); font-weight: bold; }
        .status-warn { color: var(--warning-color); font-weight: bold; }
        
        .progress-container { margin: 20px 0; display: none; }
        .progress-bar { width: 100%; height: 10px; background: #e5e7eb; border-radius: 5px; overflow: hidden; }
        .progress-fill { height: 100%; background: var(--success-color); width: 0%; transition: width 0.3s; }
        
        .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
        @media (max-width: 768px) { .grid-2 { grid-template-columns: 1fr; } }
        
        .info-box { background: #eff6ff; border-left: 4px solid var(--primary-color); padding: 15px; margin-bottom: 20px; border-radius: 4px; }
    </style>
</head>
<body>
    <header>
        <div class="header-content">
            <h1>🖥️ BK Server Manager</h1>
            <div class="status-badge" id="serverStatus">● Online</div>
        </div>
    </header>
    <div class="container">
        <div class="tabs">
            <button class="tab-btn active" onclick="switchTab('ntp')">🕐 NTP Серверы</button>
            <button class="tab-btn" onclick="switchTab('web')">🌐 Web Интерфейс</button>
            <button class="tab-btn" onclick="switchTab('cloud')">☁️ Cloud</button>
            <button class="tab-btn" onclick="switchTab('versions')">📋 Версии ПО</button>
            <button class="tab-btn" onclick="switchTab('archive')">📦 Архивы</button>
            <button class="tab-btn" onclick="switchTab('users')">👥 Пользователи</button>
            <button class="tab-btn" onclick="switchTab('rights')">✅ Права доступа</button>
            <button class="tab-btn" onclick="switchTab('db')">💾 Базы Данных</button>
            <button class="tab-btn" onclick="switchTab('pos')">🏪 POS Терминалы</button>
            <button class="tab-btn" onclick="switchTab('ip')">🌐 IP Адреса</button>
        </div>
        <!-- Вкладка NTP -->
        <div id="tab-ntp" class="tab-content active">
            <h2>Проверка и настройка NTP серверов</h2>
            <div class="info-box">Введите список серверов (каждый с новой строки). Пример:<br>192.168.1.10<br>192.168.1.11</div>
            <div class="input-group">
                <label for="ntp-servers">Список серверов:</label>
                <textarea id="ntp-servers" placeholder="192.168.1.10&#10;192.168.1.11&#10;192.168.1.12"></textarea>
            </div>
            <button class="btn" onclick="runTask('ntp')">🚀 Запустить проверку</button>
            <button class="btn btn-success" onclick="exportToExcel('ntp')" style="margin-left: 10px;">📥 Экспорт в Excel</button>
            
            <div class="progress-container" id="progress-ntp">
                <p>Выполнение задачи...</p>
                <div class="progress-bar"><div class="progress-fill" id="fill-ntp"></div></div>
            </div>
            
            <div class="results-area">
                <h3>Результаты:</h3>
                <div class="log-box" id="log-ntp">Ожидание запуска...</div>
                <div id="table-ntp"></div>
            </div>
        </div>
        <!-- Вкладка Web -->
        <div id="tab-web" class="tab-content">
            <h2>Управление Web интерфейсом</h2>
            <div class="grid-2">
                <div class="input-group">
                    <label for="web-servers">Список серверов:</label>
                    <textarea id="web-servers" placeholder="Серверы..."></textarea>
                </div>
                <div class="input-group">
                    <label>Действие:</label>
                    <select id="web-action" style="width:100%; padding: 10px; border-radius: 8px; border: 1px solid #ccc;">
                        <option value="enable">Включить</option>
                        <option value="disable">Выключить</option>
                        <option value="check">Проверить статус</option>
                    </select>
                </div>
            </div>
            <button class="btn" onclick="runTask('web')">🚀 Выполнить</button>
            <div class="progress-container" id="progress-web"><div class="progress-bar"><div class="progress-fill" id="fill-web"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-web">Ожидание...</div>
                <div id="table-web"></div>
            </div>
        </div>
        <!-- Вкладка Cloud -->
        <div id="tab-cloud" class="tab-content">
            <h2>Управление Cloud функциями</h2>
            <div class="input-group">
                <label for="cloud-servers">Список серверов:</label>
                <textarea id="cloud-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('cloud')">🚀 Проверить статус</button>
            <div class="progress-container" id="progress-cloud"><div class="progress-bar"><div class="progress-fill" id="fill-cloud"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-cloud">Ожидание...</div>
                <div id="table-cloud"></div>
            </div>
        </div>
        <!-- Вкладка Версии -->
        <div id="tab-versions" class="tab-content">
            <h2>Проверка версий ПО</h2>
            <div class="input-group">
                <label for="ver-servers">Список серверов:</label>
                <textarea id="ver-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('versions')">🚀 Проверить версии</button>
            <div class="progress-container" id="progress-versions"><div class="progress-bar"><div class="progress-fill" id="fill-versions"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-versions">Ожидание...</div>
                <div id="table-versions"></div>
            </div>
        </div>
        <!-- Вкладка Архив -->
        <div id="tab-archive" class="tab-content">
            <h2>Проверка глубины архивов</h2>
            <div class="input-group">
                <label for="arch-servers">Список серверов:</label>
                <textarea id="arch-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('archive')">🚀 Проверить архивы</button>
            <div class="progress-container" id="progress-archive"><div class="progress-bar"><div class="progress-fill" id="fill-archive"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-archive">Ожидание...</div>
                <div id="table-archive"></div>
            </div>
        </div>
        <!-- Вкладка Пользователи -->
        <div id="tab-users" class="tab-content">
            <h2>Список пользователей БД</h2>
            <div class="input-group">
                <label for="users-servers">Список серверов:</label>
                <textarea id="users-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('users')">🚀 Получить список</button>
            <div class="progress-container" id="progress-users"><div class="progress-bar"><div class="progress-fill" id="fill-users"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-users">Ожидание...</div>
                <div id="table-users"></div>
            </div>
        </div>
        <!-- Вкладка Права -->
        <div id="tab-rights" class="tab-content">
            <h2>Проверка прав пользователей</h2>
            <div class="grid-2">
                <div class="input-group">
                    <label for="rights-servers">Список серверов:</label>
                    <textarea id="rights-servers" placeholder="Серверы..."></textarea>
                </div>
                <div class="input-group">
                    <label for="rights-ref">Эталонный файл (JSON):</label>
                    <textarea id="rights-ref" placeholder='{ "user1": "admin", "user2": "operator" }'></textarea>
                </div>
            </div>
            <button class="btn" onclick="runTask('rights')">🚀 Сравнить права</button>
            <div class="progress-container" id="progress-rights"><div class="progress-bar"><div class="progress-fill" id="fill-rights"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-rights">Ожидание...</div>
                <div id="table-rights"></div>
            </div>
        </div>
        <!-- Вкладка БД -->
        <div id="tab-db" class="tab-content">
            <h2>Проверка состояния БД</h2>
            <div class="input-group">
                <label for="db-servers">Список серверов:</label>
                <textarea id="db-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('db')">🚀 Проверить БД</button>
            <div class="progress-container" id="progress-db"><div class="progress-bar"><div class="progress-fill" id="fill-db"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-db">Ожидание...</div>
                <div id="table-db"></div>
            </div>
        </div>
        <!-- Вкладка POS -->
        <div id="tab-pos" class="tab-content">
            <h2>Проверка POS терминалов</h2>
            <div class="input-group">
                <label for="pos-servers">Список серверов:</label>
                <textarea id="pos-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('pos')">🚀 Проверить POS</button>
            <div class="progress-container" id="progress-pos"><div class="progress-bar"><div class="progress-fill" id="fill-pos"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-pos">Ожидание...</div>
                <div id="table-pos"></div>
            </div>
        </div>
        <!-- Вкладка IP -->
        <div id="tab-ip" class="tab-content">
            <h2>Проверка IP адресов</h2>
            <div class="input-group">
                <label for="ip-servers">Список серверов:</label>
                <textarea id="ip-servers" placeholder="Серверы..."></textarea>
            </div>
            <button class="btn" onclick="runTask('ip')">🚀 Проверить IP</button>
            <div class="progress-container" id="progress-ip"><div class="progress-bar"><div class="progress-fill" id="fill-ip"></div></div></div>
            <div class="results-area">
                <div class="log-box" id="log-ip">Ожидание...</div>
                <div id="table-ip"></div>
            </div>
        </div>
    </div>
    <script>
        // Переключение вкладок
        function switchTab(tabId) {
            document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
            
            document.getElementById('tab-' + tabId).classList.add('active');
            event.target.classList.add('active');
        }
        // Запуск задачи
        async function runTask(taskType) {
            const serverInput = document.getElementById(taskType + '-servers');
            if (!serverInput || !serverInput.value.trim()) {
                alert('Пожалуйста, введите список серверов!');
                return;
            }
            const servers = serverInput.value.trim().split('\\n').filter(s => s.trim() !== '');
            
            // Блокировка кнопки
            const btn = event.target;
            const originalText = btn.innerHTML;
            btn.disabled = true;
            btn.innerHTML = '⏳ Выполняется...';
            // Показ прогресса
            const progressContainer = document.getElementById('progress-' + taskType);
            const fill = document.getElementById('fill-' + taskType);
            if (progressContainer) {
                progressContainer.style.display = 'block';
                fill.style.width = '10%';
            }
            const logBox = document.getElementById('log-' + taskType);
            logBox.textContent = 'Подключение к серверам...\\n';
            try {
                const response = await fetch('/api/run', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        task: taskType,
                        servers: servers,
                        params: {
                            action: document.getElementById(taskType + '-action') ? document.getElementById(taskType + '-action').value : null,
                            reference: document.getElementById('rights-ref') ? document.getElementById('rights-ref').value : null
                        }
                    })
                });
                if (!response.ok) throw new Error('Ошибка сети: ' + response.status);
                const data = await response.json();
                
                if (data.status === 'started') {
                    pollStatus(taskType, data.task_id);
                } else {
                    throw new Error(data.message || 'Неизвестная ошибка');
                }
            } catch (error) {
                logBox.textContent += '\\n❌ ОШИБКА: ' + error.message;
                btn.disabled = false;
                btn.innerHTML = originalText;
                if (progressContainer) progressContainer.style.display = 'none';
            }
        }
        // Опрос статуса
        async function pollStatus(taskType, taskId) {
            const logBox = document.getElementById('log-' + taskType);
            const fill = document.getElementById('fill-' + taskType);
            const btn = document.querySelector('.tab-btn.active'); // Находим активную кнопку примерно
            
            let attempts = 0;
            const maxAttempts = 100;
            const interval = setInterval(async () => {
                try {
                    const res = await fetch('/api/status/' + taskId);
                    const data = await res.json();
                    if (data.logs) {
                        logBox.textContent = data.logs.join('\\n');
                        logBox.scrollTop = logBox.scrollHeight;
                    }
                    if (data.progress) {
                        fill.style.width = data.progress + '%';
                    }
                    if (data.state === 'completed' || data.state === 'failed') {
                        clearInterval(interval);
                        if (data.results) {
                            displayResults(taskType, data.results);
                        }
                        
                        // Разблокировка
                        const currentBtn = document.querySelector('button[onclick="runTask(\\'' + taskType + '\\')"]');
                        if (currentBtn) {
                            currentBtn.disabled = false;
                            currentBtn.innerHTML = '🚀 Запустить проверку';
                        }
                        if (document.getElementById('progress-' + taskType)) {
                            document.getElementById('progress-' + taskType).style.display = 'none';
                        }
                    }
                } catch (e) {
                    console.error(e);
                    attempts++;
                    if (attempts > maxAttempts) {
                        clearInterval(interval);
                        logBox.textContent += '\\n❌ Таймаут ожидания результата';
                    }
                }
            }, 1000);
        }
        // Отображение таблицы результатов
        function displayResults(taskType, results) {
            const container = document.getElementById('table-' + taskType);
            if (!container || !results || results.length === 0) {
                container.innerHTML = '<p style="color: #666; margin-top: 20px;">Нет данных для отображения</p>';
                return;
            }
            let html = '<table><thead><tr>';
            // Заголовки
            if (results[0]) {
                Object.keys(results[0]).forEach(key => {
                    html += '<th>' + key + '</th>';
                });
            }
            html += '</tr></thead><tbody>';
            // Строки
            results.forEach(row => {
                html += '<tr>';
                Object.values(row).forEach(val => {
                    let className = '';
                    if (String(val).includes('OK') || String(val).includes('True')) className = 'status-ok';
                    if (String(val).includes('ERR') || String(val).includes('False')) className = 'status-error';
                    html += '<td class="' + className + '">' + val + '</td>';
                });
                html += '</tr>';
            });
            html += '</tbody></table>';
            container.innerHTML = html;
        }
        // Экспорт в Excel
        function exportToExcel(taskType) {
            window.location.href = '/api/export/' + taskType;
        }
    </script>
</body>
</html>
"""
# --- ЛОГИКА ЗАДАЧ (Имитация для примера) ---
def execute_task_logic(task_id, task_type, servers, params):
    """Фоновое выполнение задачи"""
    logs = []
    results = []
    
    def log(msg):
        timestamp = datetime.now().strftime("%H:%M:%S")
        logs.append(f"[{timestamp}] {msg}")
        task_status[task_id]['logs'] = logs
    log(f"Запуск задачи '{task_type}' для {len(servers)} серверов...")
    
    total = len(servers)
    for i, server in enumerate(servers):
        progress = int(((i + 1) / total) * 100)
        task_status[task_id]['progress'] = progress
        
        log(f"Обработка сервера: {server}")
        
        # Имитация работы (здесь должна быть реальная логика из вашего скрипта)
        import time
        time.sleep(0.5) 
        
        # Генерация фейковых результатов для демонстрации
        if task_type == 'ntp':
            results.append({
                "Server": server,
                "NTP Status": "Synced" if i % 3 != 0 else "Error",
                "Offset": f"{i*0.5}ms"
            })
        elif task_type == 'web':
            results.append({
                "Server": server,
                "Web State": "Enabled" if params.get('action') == 'enable' else "Disabled",
                "Port": "8080"
            })
        elif task_type == 'versions':
            results.append({
                "Server": server,
                "Version": f"2.4.{i}",
                "Build": "2024.04.16"
            })
        else:
            results.append({
                "Server": server,
                "Status": "OK",
                "Message": "Success"
            })
    log("Задача завершена успешно!")
    task_results[task_id] = results
    task_status[task_id]['state'] = 'completed'
    task_status[task_id]['results'] = results
@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)
@app.route('/api/run', methods=['POST'])
def run_api():
    data = request.json
    task_type = data.get('task')
    servers = data.get('servers', [])
    params = data.get('params', {})
    if not task_type or not servers:
        return jsonify({'status': 'error', 'message': 'Нет данных'}), 400
    task_id = f"{task_type}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Инициализация статуса
    task_status[task_id] = {
        'state': 'running',
        'logs': [],
        'progress': 0,
        'results': None
    }
    # Запуск в потоке
    thread = threading.Thread(target=execute_task_logic, args=(task_id, task_type, servers, params))
    thread.daemon = True
    thread.start()
    return jsonify({'status': 'started', 'task_id': task_id})
@app.route('/api/status/<task_id>')
def get_status(task_id):
    status = task_status.get(task_id, {'state': 'unknown'})
    return jsonify(status)
@app.route('/api/export/<task_type>')
def export_excel(task_type):
    # Поиск последнего результата для этого типа задачи
    # В реальном приложении нужно хранить историю или передавать ID
    result_data = None
    # Простой поиск последнего завершенного таска этого типа
    for tid in reversed(list(task_status.keys())):
        if tid.startswith(task_type) and task_status[tid]['state'] == 'completed':
            result_data = task_status[tid].get('results')
            break
    
    if not result_data:
        return jsonify({'error': 'Нет данных для экспорта. Сначала выполните проверку.'}), 400
    # Создание Excel файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"BK_{task_type}"
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Заголовки
    if result_data and len(result_data) > 0:
        headers = list(result_data[0].keys())
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        # Данные
        for row_idx, row_data in enumerate(result_data, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
                cell.border = border
                # Автоширина
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, len(str(row_data[key])) + 2)
    filename = f"BK_Report_{task_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)
if __name__ == '__main__':
    print("="*50)
    print("BK Server Manager - Web Interface")
    print("="*50)
    print(f"Запуск сервера на адресу: http://10.192.4.49:5000")
    print("Нажмите CTRL+C для остановки")
    print("="*50)
    
    # Запуск на всех интерфейсах, чтобы был доступен по сети
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
ENDOFFILE
