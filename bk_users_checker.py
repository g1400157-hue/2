#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BK Server Manager - Excel Edition (with Version-Based User Templates)
Добавлено: Система эталонов прав по версиям TRASSIR с сохранением в файл
Добавлено: Вкладка проверки статуса БД (/health -> database)
Добавлено: Вкладка проверки POS-терминалов (pos_folder2/terminals)
Добавлено: Вкладка получения IP-адреса сервера (network_interfaces/enp1s0/ip)
Добавлено: Кнопка отмены выполнения задач
Исправлено: Синтаксические ошибки в check_pos_terminal и save_excel_report
"""
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog, simpledialog
import socket
import ssl
import json
import logging
import sys
import os
import threading
import warnings
from datetime import datetime
import time
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Ошибка: Не найдена библиотека 'openpyxl'.")
    print("Выполните команду: python -m pip install openpyxl")
    sys.exit(1)
warnings.filterwarnings("ignore", category=DeprecationWarning)

# --- КОНФИГУРАЦИЯ ---
CONFIG = {
    'password': '12346',
    'ntp_expected': 'msk-v-dc01.bk.local,msk-v-dc02.bk.local',
    'web_enable_value': 1,
    'timeout': 5,
    'log_file': 'bk_manager_excel.log',
    'templates_file': 'user_templates.json'
}

if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

log_path = os.path.join(base_path, CONFIG['log_file'])
templates_path = os.path.join(base_path, CONFIG['templates_file'])

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler(log_path, encoding='utf-8', mode='a')]
)
logger = logging.getLogger(__name__)

# Цвета для Excel
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FONT_BOLD = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# Хранилище результатов по типам задач
RESULTS_STORE = {
    'ntp': [],
    'web': [],
    'cloud': [],
    'version': [],
    'archive': [],
    'users_list': [],
    'users_check': [],
    'db': [],
    'pos': [],
    'ip': []
}

# --- БАЗА ЭТАЛОНОВ ПРАВОК ---
class UserTemplatesDB:
    """Управление базой эталонов прав пользователей по версиям"""
    DEFAULT_TEMPLATES = {
        '4.3': {
            'Admin': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 1,
                    'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 1,
                    'view_button': 1, 'can_change_password': 1, 'base_rights': 75631},
            'KRU': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                    'templates_managing': 1, 'templates_sharing': 1, 'settings_button': 1, 'shutdown_button': 0, 
                    'view_button': 1, 'can_change_password': 0, 'base_rights': 1315},
            'OPS': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                    'templates_managing': 1, 'templates_sharing': 1, 'settings_button': 1, 'shutdown_button': 0, 
                    'view_button': 1, 'can_change_password': 0, 'base_rights': 291},
            'Manager': {'enable_web': 0, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                    'templates_managing': 1, 'templates_sharing': 1, 'settings_button': 1, 'shutdown_button': 0, 
                    'view_button': 0, 'can_change_password': 0, 'base_rights': 291}
        },
        '4.5': {
            'Admin': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 1,
                    'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 1,
                    'view_button': 1, 'can_change_password': 1, 'base_rights': 75631},
            'KRU': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                    'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 0,
                    'view_button': 1, 'can_change_password': 0, 'base_rights': 75043},
            'OPS': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                    'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 0,
                    'view_button': 1, 'can_change_password': 0, 'base_rights': 74019},
            'Manager': {'enable_web': 0, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                        'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 0,
                        'view_button': 0, 'can_change_password': 0, 'base_rights': 8483}
        },
        '4.6+': {
            'Admin': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 1,
                    'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 1,
                    'view_button': 1, 'can_change_password': 1, 'base_rights': 75631},
            'KRU': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                    'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 0,
                    'view_button': 1, 'can_change_password': 0, 'base_rights': 1575},
            'OPS': {'enable_web': 1, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                    'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 0,
                    'view_button': 1, 'can_change_password': 0, 'base_rights': 8995},
            'Manager': {'enable_web': 0, 'enable_local': 1, 'enable_remote': 1, 'enable_analytics': 0,
                        'templates_managing': 1, 'templates_sharing': 1, 'shutdown_button': 0,
                        'view_button': 0, 'can_change_password': 0, 'base_rights': 803}
        }
    }
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.templates = self.load()
    
    def load(self):
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                    templates = self.DEFAULT_TEMPLATES.copy()
                    templates.update(loaded)
                    return templates
            except:
                pass
        return self.DEFAULT_TEMPLATES.copy()
    
    def save(self):
        with open(self.filepath, 'w', encoding='utf-8') as f:
            json.dump(self.templates, f, indent=2, ensure_ascii=False)
    
    def get_template(self, version, user_type):
        if not version or version in ['Неизвестно', '', '-']:
            ver_key = '4.5'
        elif version.startswith('4.3') or version.startswith('4.4'):
            ver_key = '4.3'
        elif version.startswith('4.5'):
            ver_key = '4.5'
        else:
            ver_key = '4.6+'
        if ver_key in self.templates and user_type in self.templates[ver_key]:
            return self.templates[ver_key][user_type], ver_key
        return None, ver_key
    
    def save_template(self, version, user_type, settings):
        if version.startswith('4.3') or version.startswith('4.4'):
            ver_key = '4.3'
        elif version.startswith('4.5'):
            ver_key = '4.5'
        else:
            ver_key = '4.6+'
        if ver_key not in self.templates:
            self.templates[ver_key] = {}
        self.templates[ver_key][user_type] = settings
        self.save()
    
    def get_all_versions(self):
        return list(self.templates.keys())
    
    def get_all_users_for_version(self, version):
        if version in self.templates:
            return list(self.templates[version].keys())
        return []

templates_db = UserTemplatesDB(templates_path)

# --- СЕТЕВАЯ ЧАСТЬ ---
def create_no_alpn_context():
    context = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE
    try: context.set_alpn_protocols([])
    except AttributeError: pass
    context.set_ciphers('DEFAULT:@SECLEVEL=1')
    try: context.minimum_version = ssl.TLSVersion.TLSv1
    except (AttributeError, ValueError): context.options |= ssl.OP_NO_SSLv2 | ssl.OP_NO_SSLv3
    return context

def make_request(host, port, path, method='GET', data=None, sid=None):
    url_path = path
    if sid is not None:
        separator = '&' if '?' in url_path else '?'
        url_path += f"{separator}sid={sid}"
    body_content = ""
    data_exists = False
    if data is not None and len(str(data)) > 0:
        data_exists = True
    request_lines = [
        f"{method} /{url_path} HTTP/1.1",
        f"Host: {host}:{port}",
        "User-Agent: BK-Manager-Excel/1.0",
        "Accept: */*",
        "Connection: close"
    ]
    if data_exists:
        request_lines.append("Content-Type: application/x-www-form-urlencoded")
        request_lines.append(f"Content-Length: {len(str(data))}")
        body_content = str(data)
    request_str = "\r\n".join(request_lines) + "\r\n\r\n" + body_content
    context = create_no_alpn_context()
    try:
        sock = socket.create_connection((host, port), timeout=CONFIG['timeout'])
        ssock = context.wrap_socket(sock, server_hostname=host)
        ssock.sendall(request_str.encode('utf-8'))
        response_data = b""
        while True:
            chunk = ssock.recv(4096)
            if not chunk: break
            response_data += chunk
        ssock.close()
        resp_text = response_data.decode('utf-8', errors='ignore')
        if '\r\n\r\n' in resp_text:
            _, body = resp_text.split('\r\n\r\n', 1)
            body = body.strip()
            if body.startswith('{'):
                end_idx = body.rfind('}') + 1
                if end_idx > 0: return json.loads(body[:end_idx])
            return json.loads(body)
        return None
    except Exception:
        return None

def is_success(response):
    if response is None: return False
    val = response.get('success')
    if val is None: return False
    return str(val) == "1"

def parse_server_url(url_string):
    url_string = url_string.strip()
    if not url_string or url_string.startswith('#'): return None
    if '://' in url_string: url_string = url_string.split('://', 1)[1]
    if ':' in url_string:
        parts = url_string.split(':')
        host = parts[0]
        try: port = int(parts[1].split('/')[0])
        except ValueError: port = 8080
    else:
        host = url_string
        port = 8080
    return {'host': host, 'port': port, 'original': url_string}

# --- ЛОГИКА ЗАДАЧ ---
def check_ntp_single(host, port):
    path_get = f"settings/time_setup/ntp_servers?password={CONFIG['password']}"
    data = make_request(host, port, path_get)
    if data is None: return False, "Нет соединения", "", ""
    current_val = data.get('value', '')
    curr_sorted = ','.join(sorted([x.strip() for x in str(current_val).split(',')]))
    exp_sorted = ','.join(sorted([x.strip() for x in CONFIG['ntp_expected'].split(',')]))
    if curr_sorted == exp_sorted:
        return True, "Конфигурация корректна", current_val, CONFIG['ntp_expected']
    login_resp = make_request(host, port, f"login?password={CONFIG['password']}", method='POST')
    if not is_success(login_resp) or not login_resp.get('sid'):
        return False, "Ошибка входа", current_val, CONFIG['ntp_expected']
    sid = login_resp['sid']
    update_path = f"settings/time_setup/ntp_servers={CONFIG['ntp_expected']}"
    body = f"ntp_servers={CONFIG['ntp_expected']}"
    update_resp = make_request(host, port, update_path, method='POST', data=body, sid=sid)
    if is_success(update_resp):
        check_data = make_request(host, port, path_get)
        if check_data is not None:
            new_val = check_data.get('value', '')
            new_sorted = ','.join(sorted([x.strip() for x in str(new_val).split(',')]))
            if new_sorted == exp_sorted:
                return True, "Исправлено", new_val, CONFIG['ntp_expected']
    return False, "Ошибка записи", current_val, CONFIG['ntp_expected']

def check_web_status(host, port):
    path_get = f"settings/webserver/enable_webview?password={CONFIG['password']}"
    data = make_request(host, port, path_get)
    if data is None:
        return False, "Нет соединения", None
    current_val = data.get('value')
    if current_val is None:
        path_get = f"settings/webserver/enable_sdk?password={CONFIG['password']}"
        data = make_request(host, port, path_get)
        if data is None:
            return False, "Нет соединения", None
        current_val = data.get('value')
    status = "Включен" if current_val == 1 else "Выключен"
    return True, status, current_val

def set_web_status(host, port, value):
    login_resp = make_request(host, port, f"login?password={CONFIG['password']}", method='POST')
    if not is_success(login_resp) or not login_resp.get('sid'):
        return False, "Ошибка входа"
    sid = login_resp['sid']
    update_path = f"settings/webserver/enable_webview={value}"
    body = f"enable_webview={value}"
    for attempt in range(3):
        update_resp = make_request(host, port, update_path, method='POST', data=body, sid=sid)
        if is_success(update_resp):
            time.sleep(0.5)
            for check_attempt in range(3):
                time.sleep(0.3)
                check_data = make_request(host, port, f"settings/webserver/enable_webview?sid={sid}")
                if check_data is not None and check_data.get('value') == value:
                    return True, "Web включен" if value == 1 else "Web выключен"
            return True, "Web включен" if value == 1 else "Web выключен"
    return False, "Ошибка записи"

def check_cloud_status(host, port):
    path_get = f"settings/cloud/cloud_enabled?password={CONFIG['password']}"
    data = make_request(host, port, path_get)
    if data is None:
        return False, "Нет соединения", None
    current_val = data.get('value')
    if current_val is None:
        return False, "Параметр не найден", None
    status = "Включено" if current_val == 1 else "Выключено"
    return True, status, current_val

def set_cloud_status(host, port, target_value):
    path_get = f"settings/cloud/cloud_enabled?password={CONFIG['password']}"
    data = make_request(host, port, path_get)
    if data is None:
        return False, "Нет соединения"
    current_val = data.get('value')
    if current_val == target_value:
        status_msg = "Включено" if target_value == 1 else "Выключено"
        return True, f"Уже {status_msg}"
    login_resp = make_request(host, port, f"login?password={CONFIG['password']}", method='POST')
    if not is_success(login_resp) or not login_resp.get('sid'):
        return False, "Ошибка входа"
    sid = login_resp['sid']
    update_path = f"settings/cloud/cloud_enabled={target_value}"
    body = f"cloud_enabled={target_value}"
    update_resp = make_request(host, port, update_path, method='POST', data=body, sid=sid)
    if is_success(update_resp):
        time.sleep(0.5)
        check_data = make_request(host, port, path_get)
        if check_data is not None and check_data.get('value') == target_value:
            status_msg = "Включено" if target_value == 1 else "Выключено"
            return True, f"Изменено: {status_msg}"
    return False, "Ошибка записи"

def check_version_single(host, port):
    path_get = f"settings/health/product_version?password={CONFIG['password']}"
    data = make_request(host, port, path_get)
    if data is None: return False, "Нет соединения", ""
    version = data.get('value', data.get('product_version', 'Неизвестно'))
    return True, "Версия получена", str(version)

def check_archive_days(host, port, stream_type):
    if stream_type == 'main':
        path_get = f"settings/health/disks_stat_main_days?password={CONFIG['password']}"
        label = "Main"
    else:
        path_get = f"settings/health/disks_stat_subs_days?password={CONFIG['password']}"
        label = "Subs"
    data = make_request(host, port, path_get)
    if data is None:
        return False, "Нет соединения", "0"
    days_val = data.get('value', '0')
    return True, f"{label}: {days_val} дн.", str(days_val)

def check_db_status(host, port):
    """Проверка статуса БД через /health endpoint (поле database)"""
    path_get = f"health?password={CONFIG['password']}"
    data = make_request(host, port, path_get)
    if data is None:
        return False, "Нет соединения", None
    db_status = data.get('database')
    if db_status is None:
        return False, "Параметр database не найден", None
    if db_status == 1 or db_status == "1":
        return True, "БД работает", 1
    elif db_status == 0 or db_status == "0":
        return False, "Ошибка БД", 0
    else:
        return True, f"Неизвестный статус БД ({db_status})", db_status

def get_server_ip(host, port):
    path_get = f"settings/network_interfaces/enp1s0/ip?password={CONFIG['password']}"
    data = make_request(host, port, path_get)
    if data is None:
        return False, "Нет соединения", None
    ip = data.get('value')
    if ip is None:
        return False, "Параметр не найден", None
    return True, "IP получен", ip

def get_pos_terminals_list(host, port):
    path_get = f"settings/pos_folder2/terminals/?password={CONFIG['password']}"
    data = make_request(host, port, path_get)
    if data is None:
        return None
    return data.get('subdirs', [])

def check_pos_terminal(host, port, terminal_id):
    result = {'terminal_id': terminal_id, 'name': None, 'port': None, 'pos_type': None, 'pos_enable': None, 'error': None}
    
    name_data = make_request(host, port, f"settings/pos_folder2/terminals/{terminal_id}/name?password={CONFIG['password']}")
    if name_data and 'value' in name_data:
        result['name'] = name_data['value']
    
    port_data = make_request(host, port, f"settings/pos_folder2/terminals/{terminal_id}/port?password={CONFIG['password']}")
    if port_data and 'value' in port_data:
        result['port'] = port_data['value']
    
    type_data = make_request(host, port, f"settings/pos_folder2/terminals/{terminal_id}/pos_type?password={CONFIG['password']}")
    if type_data and 'value' in type_data:
        result['pos_type'] = type_data['value']
    
    enable_data = make_request(host, port, f"settings/pos_folder2/terminals/{terminal_id}/pos_enable?password={CONFIG['password']}")
    if enable_data and 'value' in enable_data:
        result['pos_enable'] = enable_data['value']
    
    if result['name'] is None or result['port'] is None or result['pos_type'] is None:
        result['error'] = "Не все параметры получены"
        return False, result
    return True, result

def get_users_list(host, port):
    resp = make_request(host, port, f"settings/users/?password={CONFIG['password']}")
    if not resp or 'subdirs' not in resp:
        return None, []
    return resp.get('subdirs', []), resp.get('values', [])

def get_user_param(host, port, user, param):
    resp = make_request(host, port, f"settings/users/{user}/{param}?password={CONFIG['password']}")
    return resp.get('value') if resp and 'value' in resp else None

def get_display_name(host, port, user):
    name = get_user_param(host, port, user, 'name')
    return name if name else user

def find_target_user(host, port, users, target):
    for u in users:
        if u.lower() == target.lower(): return u
        name = get_display_name(host, port, u)
        if name and name.lower() == target.lower(): return u
    return None

def check_user_settings(host, port, user, display_name, server_version=None):
    template, ver_key = templates_db.get_template(server_version, display_name)
    if not template:
        return None, f"Нет эталона для {display_name} (версия {ver_key})", [], [], {}
    mismatches = []
    fixes = []
    settings = {'username': user, 'name': display_name, 'version': server_version}
    for param, expected in template.items():
        actual = get_user_param(host, port, user, param)
        settings[param] = actual if actual is not None else "N/A"
        if actual is None:
            mismatches.append(f"{param}: нет данных (ожидалось {expected}) [версия {ver_key}]")
        elif str(actual) != str(expected):
            mismatches.append(f"{param}: {actual} (ожидалось {expected}) [версия {ver_key}]")
    return True, "OK", mismatches, fixes, settings

def delete_user(host, port, username):
    login_resp = make_request(host, port, f"login?password={CONFIG['password']}", method='POST')
    if not is_success(login_resp) or not login_resp.get('sid'):
        return False, "Ошибка входа"
    sid = login_resp['sid']
    resp = make_request(host, port, f"settings/users/user_add/delete_user_id={username}?sid={sid}")
    if resp and resp.get('success') in [1, '1']:
        return True, "Пользователь удален"
    return False, "Ошибка удаления"

def update_user_rights(host, port, user_guid, rights_dict):
    login_resp = make_request(host, port, f"login?password={CONFIG['password']}", method='POST')
    if not is_success(login_resp) or not login_resp.get('sid'):
        return False, "Ошибка входа"
    sid = login_resp['sid']
    for param, value in rights_dict.items():
        path = f"settings/users/{user_guid}/{param}={value}"
        resp = make_request(host, port, path, method='POST', sid=sid)
        if not is_success(resp):
            return False, f"Ошибка установки параметра {param}"
    return True, "Права обновлены"

def create_user(host, port, username, password, base_rights=0, enable_web=0, enable_local=1, enable_remote=1):
    login_resp = make_request(host, port, f"login?password={CONFIG['password']}", method='POST')
    if not is_success(login_resp) or not login_resp.get('sid'):
        return False, "Ошибка входа"
    sid = login_resp['sid']
    steps = [
        (f"settings/users/user_add/new_user_name={username}", "POST"),
        (f"settings/users/user_add/new_user_password={password}", "POST"),
        (f"settings/users/user_add/create_now=1", "POST"),
    ]
    for step_path, method in steps:
        resp = make_request(host, port, step_path, method=method, sid=sid)
        if not is_success(resp):
            return False, f"Ошибка на шаге создания: {step_path}"
    users, _ = get_users_list(host, port)
    user_guid = None
    for u in users:
        name = get_display_name(host, port, u)
        if name == username:
            user_guid = u
            break
    if not user_guid:
        return False, "Не удалось найти созданного пользователя"
    settings_to_apply = {
        'base_rights': base_rights, 'enable_web': enable_web, 'enable_local': enable_local, 'enable_remote': enable_remote,
        'can_change_password': 0, 'shutdown_button': 0, 'view_button': 0, 'templates_managing': 0, 'templates_sharing': 0
    }
    for param, value in settings_to_apply.items():
        path = f"settings/users/{user_guid}/{param}={value}"
        resp = make_request(host, port, path, method='POST', sid=sid)
        if not is_success(resp):
            return False, f"Ошибка установки параметра {param}"
    return True, "Пользователь создан"

# --- ДИАЛОГИ ---
class UserRightsDialog(tk.Toplevel):
    def __init__(self, parent, servers, templates_db):
        super().__init__(parent)
        self.title("✏️ Редактирование прав пользователей")
        self.geometry("950x750")
        self.transient(parent)
        self.grab_set()
        self.servers = servers
        self.templates_db = templates_db
        self.selected_user = None
        self.selected_server = None
        self.selected_version = None
        
        ttk.Label(self, text="Редактирование прав пользователей", font=("Segoe UI", 12, "bold")).pack(pady=10)
        version_frame = ttk.LabelFrame(self, text="Версия TRASSIR", padding="10")
        version_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(version_frame, text="Выберите версию:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.combo_version = ttk.Combobox(version_frame, values=templates_db.get_all_versions(), state="readonly", width=20)
        self.combo_version.grid(row=0, column=1, sticky=tk.W, pady=5, padx=5)
        self.combo_version.set('4.5')
        self.combo_version.bind('<<ComboboxSelected>>', self.on_version_selected)
        ttk.Button(version_frame, text="💾 Сохранить эталон", command=self.save_template).grid(row=0, column=2, pady=5, padx=5)
        ttk.Button(version_frame, text="📂 Загрузить эталон", command=self.load_template).grid(row=0, column=3, pady=5, padx=5)
        
        select_frame = ttk.LabelFrame(self, text="Выбор пользователя", padding="10")
        select_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(select_frame, text="Сервер:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.combo_server = ttk.Combobox(select_frame, width=30, state="readonly")
        self.combo_server.grid(row=0, column=1, sticky=tk.W, pady=5, padx=5)
        self.combo_server.bind('<<ComboboxSelected>>', self.on_server_selected)
        ttk.Label(select_frame, text="Пользователь:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.combo_user = ttk.Combobox(select_frame, width=30, state="readonly")
        self.combo_user.grid(row=1, column=1, sticky=tk.W, pady=5, padx=5)
        self.combo_user.bind('<<ComboboxSelected>>', self.on_user_selected)
        ttk.Button(select_frame, text="🔄 Загрузить список", command=self.load_users).grid(row=0, column=2, rowspan=2, padx=10)
        
        rights_frame = ttk.LabelFrame(self, text="Права пользователя", padding="10")
        rights_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        left_frame = ttk.Frame(rights_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        ttk.Label(left_frame, text="Разрешения на вход:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)
        self.var_local = tk.BooleanVar()
        self.var_remote = tk.BooleanVar()
        self.var_web = tk.BooleanVar()
        self.var_analytics = tk.BooleanVar()
        ttk.Checkbutton(left_frame, text="Разрешить локальный вход", variable=self.var_local).pack(anchor=tk.W)
        ttk.Checkbutton(left_frame, text="Разрешить подключение с Server/Client", variable=self.var_remote).pack(anchor=tk.W)
        ttk.Checkbutton(left_frame, text="Разрешить подключение с мобильного / из браузера", variable=self.var_web).pack(anchor=tk.W)
        ttk.Checkbutton(left_frame, text="Разрешить аналитику через сеть", variable=self.var_analytics).pack(anchor=tk.W)
        ttk.Separator(left_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        ttk.Label(left_frame, text="Ограничения интерфейса:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)
        self.var_templates_manage = tk.BooleanVar()
        self.var_templates_share = tk.BooleanVar()
        self.var_settings_btn = tk.BooleanVar()
        self.var_shutdown = tk.BooleanVar()
        self.var_view_dialog = tk.BooleanVar()
        self.var_change_password = tk.BooleanVar()
        ttk.Checkbutton(left_frame, text="Разрешить управление шаблонами", variable=self.var_templates_manage).pack(anchor=tk.W)
        ttk.Checkbutton(left_frame, text="Разрешить публикацию шаблонов в облако", variable=self.var_templates_share).pack(anchor=tk.W)
        ttk.Checkbutton(left_frame, text="Разрешить кнопку 'Настройки'", variable=self.var_settings_btn).pack(anchor=tk.W)
        ttk.Checkbutton(left_frame, text="Разрешить выключение и перезагрузку", variable=self.var_shutdown).pack(anchor=tk.W)
        ttk.Checkbutton(left_frame, text="Разрешить диалог 'Вид'", variable=self.var_view_dialog).pack(anchor=tk.W)
        ttk.Checkbutton(left_frame, text="Разрешить смену пароля", variable=self.var_change_password).pack(anchor=tk.W)
        
        right_frame = ttk.Frame(rights_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)
        ttk.Label(right_frame, text="Базовые права:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)
        self.base_rights_vars = {}
        base_rights_list = [("Просмотр", 1), ("Просмотр архива", 2), ("Слушать звук", 4), ("Просматривать видео без водяного знака", 8), ("Экспортировать архив, скриншоты", 256), ("Экспортировать архив без размытия лиц", 512), ("Редактировать закладки архива", 32), ("Использовать PTZ", 512), ("Управление", 4), ("Настройка", 8), ("Настройка пользователей и скриптов", 64)]
        for text, value in base_rights_list:
            var = tk.BooleanVar()
            self.base_rights_vars[value] = var
            ttk.Checkbutton(right_frame, text=text, variable=var).pack(anchor=tk.W)
        
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="💾 Применить изменения", command=self.apply_rights).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Закрыть", command=self.destroy).pack(side=tk.LEFT, padx=5)
        self.load_servers()

    def load_servers(self):
        server_list = [f"{s['host']}:{s['port']}" for s in self.servers]
        self.combo_server['values'] = server_list
        if server_list: self.combo_server.current(0); self.load_users()
    def load_users(self):
        if not self.combo_server.get(): return
        host, port = self.combo_server.get().split(':')
        port = int(port)
        users, _ = get_users_list(host, port)
        if users:
            self.combo_user['values'] = [f"{get_display_name(host, port, u)} ({u})" for u in users]
            if self.combo_user['values']: self.combo_user.current(0); self.on_user_selected()
    def on_version_selected(self, event=None): self.selected_version = self.combo_version.get(); self.load_template_values()
    def on_server_selected(self, event=None): self.load_users()
    def on_user_selected(self, event=None):
        if not self.combo_user.get(): return
        user_str = self.combo_user.get()
        name = user_str.split(' (')[0] if ' (' in user_str else user_str
        guid = user_str.split('(')[1].rstrip(')') if ' (' in user_str else user_str
        host, port = self.combo_server.get().split(':')
        self.selected_user, self.selected_server = guid, (host, int(port))
        self.load_current_rights(host, int(port), guid)
    def load_template_values(self):
        version = self.combo_version.get()
        user_type = self.combo_user.get().split(' (')[0] if ' (' in self.combo_user.get() else self.combo_user.get()
        template, _ = self.templates_db.get_template(version, user_type)
        if template:
            self.var_web.set(template.get('enable_web', 0) == 1)
            self.var_local.set(template.get('enable_local', 0) == 1)
            self.var_remote.set(template.get('enable_remote', 0) == 1)
            self.var_analytics.set(template.get('enable_analytics', 0) == 1)
            self.var_templates_manage.set(template.get('templates_managing', 0) == 1)
            self.var_templates_share.set(template.get('templates_sharing', 0) == 1)
            self.var_settings_btn.set(template.get('settings_button', 0) == 1)
            self.var_shutdown.set(template.get('shutdown_button', 0) == 1)
            self.var_view_dialog.set(template.get('view_button', 0) == 1)
            self.var_change_password.set(template.get('can_change_password', 0) == 1)
            base_rights = template.get('base_rights', 0)
            for value, var in self.base_rights_vars.items(): var.set((base_rights & value) != 0)
    def load_current_rights(self, host, port, user_guid):
        for p, v in [('enable_local', self.var_local), ('enable_remote', self.var_remote), ('enable_web', self.var_web), 
                     ('enable_analytics', self.var_analytics), ('templates_managing', self.var_templates_manage), 
                     ('templates_sharing', self.var_templates_share), ('settings_button', self.var_settings_btn), 
                     ('shutdown_button', self.var_shutdown), ('view_button', self.var_view_dialog), ('can_change_password', self.var_change_password)]:
            val = get_user_param(host, port, user_guid, p)
            v.set(val == '1')
        base_rights = get_user_param(host, port, user_guid, 'base_rights')
        if base_rights:
            try:
                rv = int(base_rights)
                for value, var in self.base_rights_vars.items(): var.set((rv & value) != 0)
            except ValueError: pass
    def save_template(self):
        version = self.combo_version.get()
        user_type = self.combo_user.get().split(' (')[0] if ' (' in self.combo_user.get() else self.combo_user.get()
        base_rights = sum(v for v, var in self.base_rights_vars.items() if var.get())
        settings = {'enable_web': 1 if self.var_web.get() else 0, 'enable_local': 1 if self.var_local.get() else 0,
                    'enable_remote': 1 if self.var_remote.get() else 0, 'enable_analytics': 1 if self.var_analytics.get() else 0,
                    'templates_managing': 1 if self.var_templates_manage.get() else 0, 'templates_sharing': 1 if self.var_templates_share.get() else 0,
                    'settings_button': 1 if self.var_settings_btn.get() else 0, 'shutdown_button': 1 if self.var_shutdown.get() else 0,
                    'view_button': 1 if self.var_view_dialog.get() else 0, 'can_change_password': 1 if self.var_change_password.get() else 0, 'base_rights': base_rights}
        self.templates_db.save_template(version, user_type, settings)
        messagebox.showinfo("Успех", f"Эталон сохранен для версии {version}!\nФайл: {templates_path}")
    def load_template(self):
        if os.path.exists(templates_path):
            self.templates_db.templates = self.templates_db.load()
            self.combo_version['values'] = self.templates_db.get_all_versions()
            messagebox.showinfo("Успех", f"Эталоны загружены из файла:\n{templates_path}")
        else: messagebox.showwarning("Внимание", "Файл эталонов не найден!")
    def apply_rights(self):
        if not self.selected_user or not self.selected_server: return messagebox.showerror("Ошибка", "Выберите пользователя!")
        host, port = self.selected_server
        base_rights = sum(v for v, var in self.base_rights_vars.items() if var.get())
        rights = {'enable_local': 1 if self.var_local.get() else 0, 'enable_remote': 1 if self.var_remote.get() else 0,
                  'enable_web': 1 if self.var_web.get() else 0, 'enable_analytics': 1 if self.var_analytics.get() else 0,
                  'templates_managing': 1 if self.var_templates_manage.get() else 0, 'templates_sharing': 1 if self.var_templates_share.get() else 0,
                  'settings_button': 1 if self.var_settings_btn.get() else 0, 'shutdown_button': 1 if self.var_shutdown.get() else 0,
                  'view_button': 1 if self.var_view_dialog.get() else 0, 'can_change_password': 1 if self.var_change_password.get() else 0, 'base_rights': base_rights}
        ok, msg = update_user_rights(host, port, self.selected_user, rights)
        if ok: messagebox.showinfo("Успех", "Права пользователя обновлены!"); self.load_current_rights(host, port, self.selected_user)
        else: messagebox.showerror("Ошибка", f"Не удалось обновить права:\n{msg}")

class UserDeleteDialog(tk.Toplevel):
    def __init__(self, parent, servers):
        super().__init__(parent)
        self.title("🗑 Удаление пользователей")
        self.geometry("900x600")
        self.transient(parent)
        self.grab_set()
        self.servers = servers
        self.users_data = {}
        ttk.Label(self, text="Выберите пользователей для удаления:", font=("Segoe UI", 11, "bold")).pack(pady=10)
        columns = ('server', 'username', 'name', 'web', 'local', 'remote')
        self.tree = ttk.Treeview(self, columns=columns, show='headings', height=15)
        for col, w in zip(columns, [150, 100, 120, 50, 50, 50]): self.tree.column(col, width=w); self.tree.heading(col, text=col.capitalize())
        scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10)
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="🔄 Обновить", command=self.load_users).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="✅ Выделить всех", command=self.select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Снять выделение", command=self.deselect_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑 Удалить выбранных", command=self.delete_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Закрыть", command=self.destroy).pack(side=tk.LEFT, padx=5)
        self.load_users()
    def load_users(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        self.users_data = {}
        for srv in self.servers:
            host, port = srv['host'], srv['port']
            try:
                users, _ = get_users_list(host, port)
                if users:
                    self.users_data[host] = {}
                    for user_guid in users:
                        name = get_display_name(host, port, user_guid)
                        w, l, r = get_user_param(host, port, user_guid, 'enable_web'), get_user_param(host, port, user_guid, 'enable_local'), get_user_param(host, port, user_guid, 'enable_remote')
                        self.users_data[host][user_guid] = {'name': name, 'web': w, 'local': l, 'remote': r}
                        self.tree.insert('', tk.END, values=(host, user_guid, name, '✓' if w=='1' else '', '✓' if l=='1' else '', '✓' if r=='1' else ''))
            except Exception as e: self.tree.insert('', tk.END, values=(host, "Ошибка", str(e), "", "", ""))
    def select_all(self):
        for item in self.tree.get_children(): self.tree.selection_add(item)
    def deselect_all(self): self.tree.selection_remove(self.tree.get_children())
    def delete_selected(self):
        selected = self.tree.selection()
        if not selected: return messagebox.showwarning("Внимание", "Выберите пользователя для удаления!")
        if not messagebox.askyesno("Подтверждение", f"Удалить {len(selected)} пользователей?"): return
        res = {'success': 0, 'error': 0}
        for item in selected:
            srv, usr = self.tree.item(item)['values'][:2]
            for s in self.servers:
                if s['host'] == srv:
                    ok, _ = delete_user(s['host'], s['port'], usr)
                    if ok: res['success'] += 1; self.tree.delete(item)
                    else: res['error'] += 1
                    break
        messagebox.showinfo("Результат", f"Удалено: {res['success']}\nОшибок: {res['error']}")

class UserCreateDialog(tk.Toplevel):
    def __init__(self, parent, servers):
        super().__init__(parent)
        self.title("➕ Создание пользователей")
        self.geometry("600x500")
        self.transient(parent)
        self.grab_set()
        self.servers = servers
        ttk.Label(self, text="Создание нового пользователя", font=("Segoe UI", 12, "bold")).pack(pady=10)
        form = ttk.LabelFrame(self, text="Параметры пользователя", padding="10")
        form.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(form, text="Имя:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.entry_username = ttk.Entry(form, width=30); self.entry_username.grid(row=0, column=1, sticky=tk.W, pady=5)
        ttk.Label(form, text="Пароль:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.entry_password = ttk.Entry(form, width=30, show="*"); self.entry_password.grid(row=1, column=1, sticky=tk.W, pady=5)
        rights = ttk.LabelFrame(self, text="Права доступа", padding="10")
        rights.pack(fill=tk.X, padx=10, pady=5)
        self.var_web = tk.BooleanVar()
        self.var_local = tk.BooleanVar(value=True)
        self.var_remote = tk.BooleanVar(value=True)
        self.var_analytics = tk.BooleanVar()
        ttk.Checkbutton(rights, text="Web доступ", variable=self.var_web).grid(row=0, column=0, sticky=tk.W)
        ttk.Checkbutton(rights, text="Локальный вход", variable=self.var_local).grid(row=0, column=1, sticky=tk.W)
        ttk.Checkbutton(rights, text="Удаленный доступ", variable=self.var_remote).grid(row=0, column=2, sticky=tk.W)
        ttk.Checkbutton(rights, text="Аналитика", variable=self.var_analytics).grid(row=1, column=0, sticky=tk.W)
        ttk.Label(self, text="Пользователь будет создан на всех серверах из списка:", font=("Segoe UI", 9, "italic")).pack(pady=5)
        btns = ttk.Frame(self); btns.pack(pady=10)
        ttk.Button(btns, text="✅ Создать", command=self.create_user).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="❌ Отмена", command=self.destroy).pack(side=tk.LEFT, padx=5)
    def create_user(self):
        u, p = self.entry_username.get().strip(), self.entry_password.get().strip()
        if not u or not p: return messagebox.showerror("Ошибка", "Введите имя и пароль!")
        if not messagebox.askyesno("Подтверждение", f"Создать '{u}' на {len(self.servers)} серверах?"): return
        res = {'success': 0, 'error': 0}
        base = 1 if self.var_analytics.get() else 0
        for s in self.servers:
            ok, _ = create_user(s['host'], s['port'], u, p, base_rights=base, enable_web=1 if self.var_web.get() else 0, enable_local=1 if self.var_local.get() else 0, enable_remote=1 if self.var_remote.get() else 0)
            if ok: res['success'] += 1
            else: res['error'] += 1
        messagebox.showinfo("Результат", f"Создано: {res['success']}\nОшибок: {res['error']}")
        self.destroy()

class NTPServersDialog(tk.Toplevel):
    def __init__(self, parent, current_ntp):
        super().__init__(parent)
        self.title("Настройка NTP серверов")
        self.geometry("500x400")
        self.transient(parent)
        self.grab_set()
        self.ntp_servers = [s.strip() for s in current_ntp.split(',') if s.strip()] if current_ntp else []
        ttk.Label(self, text="Список NTP серверов:", font=("Segoe UI", 10, "bold")).pack(pady=10)
        lf = ttk.Frame(self); lf.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.listbox = tk.Listbox(lf, font=("Consolas", 9), height=10)
        sb = ttk.Scrollbar(lf, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=sb.set)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True); sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.update_listbox()
        bf = ttk.Frame(self); bf.pack(pady=5)
        ttk.Button(bf, text="+ Добавить", command=self.add_server).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="- Удалить", command=self.remove_server).pack(side=tk.LEFT, padx=5)
        ocf = ttk.Frame(self); ocf.pack(pady=10)
        ttk.Button(ocf, text="OK", command=self.on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(ocf, text="Отмена", command=self.destroy).pack(side=tk.LEFT, padx=5)
    def update_listbox(self):
        self.listbox.delete(0, tk.END)
        for i, s in enumerate(self.ntp_servers, 1): self.listbox.insert(tk.END, f"{i}. {s}")
    def add_server(self):
        s = simpledialog.askstring("Добавить NTP сервер", "Введите адрес NTP сервера:")
        if s and s.strip(): self.ntp_servers.append(s.strip()); self.update_listbox()
    def remove_server(self):
        sel = self.listbox.curselection()
        if sel: del self.ntp_servers[sel[0]]; self.update_listbox()
    def on_ok(self): self.destroy()

# --- GUI КЛАСС ---
class BKManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BK Server Manager")
        self.root.geometry("1200x800")
        self.root.minsize(900, 600)
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TButton", padding=6, relief="flat", background="#ccc", font=("Segoe UI", 10, "bold"))
        style.configure("TLabel", font=("Segoe UI", 10))
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        input_frame = ttk.LabelFrame(main_frame, text="Список серверов (URL)", padding="5")
        input_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        tb = ttk.Frame(input_frame); tb.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(tb, text="📋 Вставить", command=self.paste_servers).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="🗑 Очистить", command=self.clear_input).pack(side=tk.LEFT, padx=2)
        self.lbl_count = ttk.Label(tb, text="Серверов: 0", foreground="blue"); self.lbl_count.pack(side=tk.RIGHT)
        self.text_input = scrolledtext.ScrolledText(input_frame, height=8, font=("Consolas", 9))
        self.text_input.pack(fill=tk.BOTH, expand=True)
        self.text_input.bind('<KeyRelease>', self.update_count)
        self.text_input.insert('1.0', "# Вставьте список серверов здесь:\n# https://BKV0001-1:8080")
        self.text_input.bind('<FocusIn>', self.clear_placeholder)
        
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=10)
        for name, func in [("⏰ NTP", self.setup_ntp_tab), ("🌐 Web", self.setup_web_tab), ("☁️ Облако", self.setup_cloud_tab), 
                           ("📦 Версия ПО", self.setup_version_tab), ("💾 Архив", self.setup_archive_tab), ("🗄️ База Данных", self.setup_db_tab), 
                           ("🧾 POS Терминалы", self.setup_pos_tab), ("🌐 IP Адрес", self.setup_ip_tab), ("👥 Пользователи", self.setup_users_tab)]:
            f = ttk.Frame(self.notebook); self.notebook.add(f, text=name); func(f)
            
        style.configure("green.Horizontal.TProgressbar", troughcolor='#dddddd', background='#28a745', lightcolor='#28a745', darkcolor='#28a745', bordercolor='#28a745')
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100, style="green.Horizontal.TProgressbar")
        self.progress_bar.pack(fill=tk.X, pady=5)
        self.status_label = ttk.Label(main_frame, text="Готов к работе", anchor="w", font=("Segoe UI", 10, "bold"))
        self.status_label.pack(fill=tk.X)
        
        log_frame = ttk.LabelFrame(main_frame, text="Результаты выполнения", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.text_log = scrolledtext.ScrolledText(log_frame, height=15, font=("Consolas", 9), state='disabled')
        self.text_log.pack(fill=tk.BOTH, expand=True)
        for tag, color in [("info", "blue"), ("success", "green"), ("error", "red"), ("warning", "orange"), ("header", "black"), ("user", "purple"), ("fix", "darkgreen")]:
            self.text_log.tag_configure(tag, foreground=color, font=("Consolas", 10 if tag=="header" else 9, "bold" if tag in ["header","user","fix"] else "normal"))
        
        self.is_running = False
        self.cancel_requested = False
        self.current_log_text = ""
        
        cp = ttk.Frame(main_frame); cp.pack(fill=tk.X, pady=5)
        self.btn_copy_log = ttk.Button(cp, text="📋 Копировать", command=self.copy_log); self.btn_copy_log.pack(side=tk.LEFT, padx=5)
        self.btn_save_excel = ttk.Button(cp, text="💾 Excel Отчет", command=self.save_excel_report); self.btn_save_excel.pack(side=tk.LEFT, padx=5)
        self.btn_clear_log = ttk.Button(cp, text="🗑 Очистить", command=self.clear_log); self.btn_clear_log.pack(side=tk.LEFT, padx=5)
        self.btn_cancel = ttk.Button(cp, text="❌ Отменить", command=self.cancel_task, state='disabled'); self.btn_cancel.pack(side=tk.LEFT, padx=5)

    def cancel_task(self):
        if messagebox.askyesno("Подтверждение", "Отменить выполнение задачи?"):
            self.cancel_requested = True
            self.status_label.config(text="⏹ Отмена задачи...")
            self.btn_cancel.config(state='disabled')

    def setup_ntp_tab(self, frame):
        ttk.Label(frame, text="Проверка и настройка NTP серверов", font=("Segoe UI", 11, "bold")).pack(pady=10)
        ttk.Button(frame, text="⚙️ Настроить NTP", command=self.open_ntp_settings).pack(pady=5)
        self.ntp_label = ttk.Label(frame, text=f"Текущие: {CONFIG['ntp_expected']}", font=("Consolas", 9), foreground="blue"); self.ntp_label.pack(pady=5)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_ntp_check = ttk.Button(bf, text="🔍 Проверить", command=lambda: self.start_task('ntp_check')); self.btn_ntp_check.pack(side=tk.LEFT, padx=5)
        self.btn_ntp_fix = ttk.Button(bf, text="🔧 Исправить", command=lambda: self.start_task('ntp_fix')); self.btn_ntp_fix.pack(side=tk.LEFT, padx=5)
    def setup_web_tab(self, frame):
        ttk.Label(frame, text="Управление Web-интерфейсом", font=("Segoe UI", 11, "bold")).pack(pady=10)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_web_status = ttk.Button(bf, text="🔍 Статус", command=lambda: self.start_task('web_status')); self.btn_web_status.pack(side=tk.LEFT, padx=5)
        self.btn_web_enable = ttk.Button(bf, text="✅ Включить", command=lambda: self.start_task('web_enable')); self.btn_web_enable.pack(side=tk.LEFT, padx=5)
        self.btn_web_disable = ttk.Button(bf, text="❌ Выключить", command=lambda: self.start_task('web_disable')); self.btn_web_disable.pack(side=tk.LEFT, padx=5)
    def setup_cloud_tab(self, frame):
        ttk.Label(frame, text="Управление подключением к облаку", font=("Segoe UI", 11, "bold")).pack(pady=10)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_cloud_status = ttk.Button(bf, text="🔍 Статус", command=lambda: self.start_task('cloud_status')); self.btn_cloud_status.pack(side=tk.LEFT, padx=5)
        self.btn_cloud_enable = ttk.Button(bf, text="☁️ Включить", command=lambda: self.start_task('cloud_enable')); self.btn_cloud_enable.pack(side=tk.LEFT, padx=5)
        self.btn_cloud_disable = ttk.Button(bf, text="🚫 Выключить", command=lambda: self.start_task('cloud_disable')); self.btn_cloud_disable.pack(side=tk.LEFT, padx=5)
    def setup_version_tab(self, frame):
        ttk.Label(frame, text="Получение версии ПО TRASSIR", font=("Segoe UI", 11, "bold")).pack(pady=10)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_version = ttk.Button(bf, text="📦 Получить версию", command=lambda: self.start_task('version')); self.btn_version.pack(side=tk.LEFT, padx=5)
    def setup_archive_tab(self, frame):
        ttk.Label(frame, text="Проверка глубины архива", font=("Segoe UI", 11, "bold")).pack(pady=10)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_archive_main = ttk.Button(bf, text="📊 Main", command=lambda: self.start_task('archive_main')); self.btn_archive_main.pack(side=tk.LEFT, padx=5)
        self.btn_archive_subs = ttk.Button(bf, text="📊 Subs", command=lambda: self.start_task('archive_subs')); self.btn_archive_subs.pack(side=tk.LEFT, padx=5)
        self.btn_archive_both = ttk.Button(bf, text="📊 Оба", command=lambda: self.start_task('archive_both')); self.btn_archive_both.pack(side=tk.LEFT, padx=5)
    def setup_db_tab(self, frame):
        ttk.Label(frame, text="Проверка статуса Базы Данных", font=("Segoe UI", 11, "bold")).pack(pady=10)
        ttk.Label(frame, text="Параметр: /health → database\n1 = Работает, 0 = Ошибка БД", font=("Consolas", 9), foreground="gray").pack(pady=5)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_db_check = ttk.Button(bf, text="🗄️ Проверить БД", command=lambda: self.start_task('db_check')); self.btn_db_check.pack(side=tk.LEFT, padx=5)
    def setup_pos_tab(self, frame):
        ttk.Label(frame, text="Проверка POS-терминалов", font=("Segoe UI", 11, "bold")).pack(pady=10)
        ttk.Label(frame, text="Проверяет параметры терминалов:\n• Имя (name) • Порт (port) • Протокол (pos_type) • Статус (pos_enable)", font=("Consolas", 9), foreground="gray", justify=tk.LEFT).pack(pady=5)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_pos_check = ttk.Button(bf, text="🧾 Проверить POS", command=lambda: self.start_task('pos_check')); self.btn_pos_check.pack(side=tk.LEFT, padx=5)
    def setup_ip_tab(self, frame):
        ttk.Label(frame, text="Получение IP-адреса сервера", font=("Segoe UI", 11, "bold")).pack(pady=10)
        ttk.Label(frame, text="Параметр: settings/network_interfaces/enp1s0/ip", font=("Consolas", 9), foreground="gray").pack(pady=5)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_ip_check = ttk.Button(bf, text="🌐 Получить IP", command=lambda: self.start_task('ip_check')); self.btn_ip_check.pack(side=tk.LEFT, padx=5)
    def setup_users_tab(self, frame):
        ttk.Label(frame, text="Просмотр и управление пользователями", font=("Segoe UI", 11, "bold")).pack(pady=10)
        bf = ttk.Frame(frame); bf.pack(pady=10)
        self.btn_users_list = ttk.Button(bf, text="👥 Список", command=lambda: self.start_task('users_list')); self.btn_users_list.pack(side=tk.LEFT, padx=5)
        self.btn_users_check = ttk.Button(bf, text="🔍 Права", command=lambda: self.start_task('users_check')); self.btn_users_check.pack(side=tk.LEFT, padx=5)
        self.btn_users_edit = ttk.Button(bf, text="✏️ Редактировать права", command=self.open_user_rights); self.btn_users_edit.pack(side=tk.LEFT, padx=5)
        self.btn_users_delete = ttk.Button(bf, text="🗑 Удалить", command=self.open_user_delete); self.btn_users_delete.pack(side=tk.LEFT, padx=5)
        self.btn_users_create = ttk.Button(bf, text="➕ Создать", command=self.open_user_create); self.btn_users_create.pack(side=tk.LEFT, padx=5)

    def open_user_rights(self):
        s = self.get_servers()
        if not s: return messagebox.showerror("Ошибка", "Список серверов пуст!")
        UserRightsDialog(self.root, s, templates_db)
    def open_user_delete(self):
        s = self.get_servers()
        if not s: return messagebox.showerror("Ошибка", "Список серверов пуст!")
        UserDeleteDialog(self.root, s)
    def open_user_create(self):
        s = self.get_servers()
        if not s: return messagebox.showerror("Ошибка", "Список серверов пуст!")
        UserCreateDialog(self.root, s)
    def open_ntp_settings(self):
        d = NTPServersDialog(self.root, CONFIG['ntp_expected'])
        self.root.wait_window(d)
        if d.ntp_servers:
            CONFIG['ntp_expected'] = ','.join(d.ntp_servers)
            self.ntp_label.config(text=f"Текущие: {CONFIG['ntp_expected']}")
            self.log(f"✅ NTP обновлены: {CONFIG['ntp_expected']}", "info")
    def clear_placeholder(self, event):
        if self.text_input.get("1.0", tk.END).strip().startswith("# Вставьте список"):
            self.text_input.delete("1.0", tk.END); self.update_count()
    def update_count(self, event=None):
        self.lbl_count.config(text=f"Серверов: {sum(1 for l in self.text_input.get('1.0', tk.END).splitlines() if parse_server_url(l))}")
    def clear_input(self): self.text_input.delete("1.0", tk.END); self.update_count()
    def paste_servers(self):
        try:
            cb = self.root.clipboard_get()
            if cb:
                if self.text_input.get("1.0", tk.END).strip().startswith("# Вставьте список"): self.text_input.delete("1.0", tk.END)
                self.text_input.insert(tk.END, cb); self.update_count(); self.log("✅ Список вставлен.", "info")
        except Exception as e: messagebox.showerror("Ошибка", f"Не удалось вставить: {e}")
    def log(self, msg, tag=None):
        self.text_log.config(state='normal'); self.text_log.insert(tk.END, msg + "\n", tag); self.text_log.see(tk.END); self.text_log.config(state='disabled')
        self.current_log_text += msg + "\n"
    def clear_log(self):
        self.text_log.config(state='normal'); self.text_log.delete("1.0", tk.END); self.text_log.config(state='disabled')
        self.current_log_text = ""
        for k in RESULTS_STORE: RESULTS_STORE[k] = []
        self.log("--- Лог очищен ---", "info")
    def copy_log(self):
        if not self.current_log_text: return messagebox.showinfo("Инфо", "Лог пуст.")
        self.root.clipboard_clear(); self.root.clipboard_append(self.current_log_text)
        messagebox.showinfo("Успех", "Отчет скопирован!")
    
    def save_excel_report(self):
        has_data = any(len(v) > 0 for v in RESULTS_STORE.values())
        if not has_data:
            messagebox.showwarning("Внимание", "Нет данных для сохранения!\nСначала выполните проверки.")
            return
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        fp = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"BK_Full_Report_{ts}.xlsx", title="Сохранить отчет Excel")
        if not fp: return
        try:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames: del wb['Sheet']
            def create_sheet(name, headers, data):
                if not data: return
                ws = wb.create_sheet(title=name)
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(color="FFFFFF", bold=True); cell.alignment = CENTER_ALIGN
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                for row in data:
                    pr = []
                    for idx, val in enumerate(row):
                        if name == 'Archive' and idx in [2, 3]:
                            try: pr.append(int(float(val)) if val and val != "-" else None)
                            except: pr.append(val)
                        else: pr.append(val)
                    ws.append(pr)
                for col in ws.columns:
                    ml = max(len(str(c.value)) for c in col if c.value)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 50)
                ws.auto_filter.ref = ws.dimensions
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
                    sv = str(row[0].value).upper() if row[0].value else ""
                    fc = FILL_GREEN if any(x in sv for x in ["OK","CORRECT","ВКЛ","ВЫКЛ","РАБОТАЕТ"]) else (FILL_YELLOW if "WARN" in sv else (FILL_RED if any(x in sv for x in ["ERR","НЕТ","ОШИБ","ERROR"]) else None))
                    if fc:
                        for c in row: c.fill = fc; c.alignment = CENTER_ALIGN if c.column <= 2 else Alignment(horizontal="left", vertical="center")
            
            create_sheet("NTP", ["Статус", "Сервер", "Сообщение", "NTP (тек)", "NTP (ожид)", "Время"], RESULTS_STORE['ntp'])
            create_sheet("Web", ["Статус", "Сервер", "Сообщение", "Время"], RESULTS_STORE['web'])
            create_sheet("Cloud", ["Статус", "Сервер", "Сообщение", "Время"], RESULTS_STORE['cloud'])
            create_sheet("Version", ["Статус", "Сервер", "Версия ПО", "Время"], RESULTS_STORE['version'])
            create_sheet("Archive", ["Статус", "Сервер", "Архив Main (дн)", "Архив Subs (дн)", "Время"], RESULTS_STORE['archive'])
            create_sheet("Database", ["Статус", "Сервер", "Сообщение", "Код статуса", "Время"], RESULTS_STORE['db'])
            
            if RESULTS_STORE['pos']:
                ws_pos = wb.create_sheet(title="POS_Terminals")
                max_t = max(len(r[3].split(' | ')) for r in RESULTS_STORE['pos'] if len(r)>=4 and r[3] and r[3]!='-') if any(len(r)>=4 and r[3] and r[3]!='-' for r in RESULTS_STORE['pos']) else 0
                ws_pos.append(["Статус", "Сервер", "Время"] + [f"Терминал {i} (ID:Имя:Порт:Статус)" for i in range(1, max_t+1)])
                for cell in ws_pos[1]: cell.font = Font(color="FFFFFF", bold=True); cell.alignment = CENTER_ALIGN; cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                for r in RESULTS_STORE['pos']:
                    st, srv, ts = r[0], r[1], r[4] if len(r)>4 else ""
                    dets = r[3].split(' | ') if len(r)>3 and r[3]!='-' else []
                    ws_pos.append([st, srv, ts] + dets + [""]*(max_t-len(dets)))
                for col in ws_pos.columns:
                    ml = max(len(str(c.value)) for c in col if c.value)
                    ws_pos.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 50)
                ws_pos.auto_filter.ref = ws_pos.dimensions
                for row in ws_pos.iter_rows(min_row=2, max_row=ws_pos.max_row, values_only=False):
                    sv = str(row[0].value).upper() if row[0].value else ""
                    fc = FILL_GREEN if "OK" in sv else (FILL_RED if "ERROR" in sv else None)
                    if fc:
                        for c in row: c.fill = fc; c.alignment = CENTER_ALIGN if c.column <= 3 else Alignment(horizontal="left", vertical="center")
            
            create_sheet("IP_Addresses", ["Статус", "Сервер", "IP-адрес", "Время"], RESULTS_STORE['ip'])
            create_sheet("Users_List", ["Сервер", "Количество", "Список пользователей", "Время"], RESULTS_STORE['users_list'])
            create_sheet("Users_Check", ["Сервер", "Пользователь", "GUID", "Статус", "Расхождения", "Время"], RESULTS_STORE['users_check'])
            
            wb.save(fp)
            messagebox.showinfo("Успех", f"Отчет сохранен:\n{fp}\nСоздано листов: {len(wb.sheetnames)}")
            logger.info(f"Excel отчет сохранен в {fp}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить Excel:\n{e}")
            logger.exception("Ошибка сохранения Excel")

    def start_task(self, task_type):
        if self.is_running: return messagebox.showwarning("Внимание", "Задача уже выполняется!")
        servers = self.get_servers()
        if not servers: return messagebox.showerror("Ошибка", "Список серверов пуст!")
        names = {'ntp_check': 'Проверка NTP', 'ntp_fix': 'Исправление NTP', 'web_status': 'Статус Web', 'web_enable': 'Включение Web', 'web_disable': 'Отключение Web',
                 'cloud_status': 'Статус Облака', 'cloud_enable': 'Включение облака', 'cloud_disable': 'Отключение облака', 'version': 'Версия ПО',
                 'archive_main': 'Архив (Main)', 'archive_subs': 'Архив (Subs)', 'archive_both': 'Архив (Все)', 'db_check': 'Проверка БД',
                 'pos_check': 'Проверка POS', 'ip_check': 'Получение IP', 'users_list': 'Список пользователей', 'users_check': 'Проверка прав'}
        if not messagebox.askyesno("Подтверждение", f"Запустить задачу '{names.get(task_type, task_type)}' для {len(servers)} серверов?"): return
        self.is_running, self.cancel_requested = True, False
        for b in [self.btn_ntp_check, self.btn_ntp_fix, self.btn_web_status, self.btn_web_enable, self.btn_web_disable, self.btn_cloud_status, self.btn_cloud_enable, self.btn_cloud_disable, self.btn_version, self.btn_archive_main, self.btn_archive_subs, self.btn_archive_both, self.btn_db_check, self.btn_pos_check, self.btn_ip_check, self.btn_users_list, self.btn_users_check, self.btn_copy_log, self.btn_save_excel, self.btn_clear_log]:
            try: b.config(state='disabled')
            except: pass
        self.btn_cancel.config(state='normal')
        self.progress_var.set(0); self.clear_log()
        threading.Thread(target=self.run_process, args=(task_type, servers), daemon=True).start()

    def run_process(self, task_type, servers):
        total = len(servers)
        rc = {'ok': 0, 'err': 0}
        names = {'ntp_check': 'Проверка NTP', 'ntp_fix': 'Исправление NTP', 'web_status': 'Статус Web', 'web_enable': 'Включение Web', 'web_disable': 'Отключение Web',
                 'cloud_status': 'Статус Облака', 'cloud_enable': 'Включение облака', 'cloud_disable': 'Отключение облака', 'version': 'Версия ПО',
                 'archive_main': 'Архив (Main)', 'archive_subs': 'Архив (Subs)', 'archive_both': 'Архив (Все)', 'db_check': 'Проверка БД',
                 'pos_check': 'Проверка POS', 'ip_check': 'Получение IP', 'users_list': 'Список пользователей', 'users_check': 'Проверка прав'}
        dn = "Облако (ВКЛ)" if task_type=='cloud_enable' else ("Облако (ВЫКЛ)" if task_type=='cloud_disable' else names.get(task_type, task_type).upper())
        self.root.after(0, lambda: self.log(f"=== ЗАПУСК: {dn} ===\nВсего серверов: {total}\nВремя: {datetime.now().strftime('%H:%M:%S')}\n", "header"))
        for i, srv in enumerate(servers):
            if self.cancel_requested:
                self.root.after(0, lambda: self.log(f"\n=== ЗАДАЧА ОТМЕНЕНА ПОЛЬЗОВАТЕЛЕМ ===\nОбработано серверов: {i}/{total}", "warning"))
                break
            host, port, sn = srv['host'], srv['port'], srv['host']
            self.root.after(0, lambda s=sn: self.status_label.config(text=f"Обработка: {s}"))
            es, em = "OK", "Проверка пройдена"
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if task_type in ['ntp_check', 'ntp_fix']:
                ok, msg, cur, exp = check_ntp_single(host, port)
                rc['ok' if ok else 'err'] += 1
                if not ok: es, em = "ERROR", msg
                self.root.after(0, lambda h=host, m=msg: self.log(f"[{i+1}/{total}] {h}: {m}", "success" if ok else "error"))
                RESULTS_STORE['ntp'].append([es, sn, msg, cur or "-", exp, ts])
            elif task_type == 'web_status':
                ok, st, _ = check_web_status(host, port)
                em = f"Web {st}"
                rc['ok' if ok else 'err'] += 1
                if not ok: es = "ERROR"
                self.root.after(0, lambda h=host, s=st: self.log(f"[{i+1}/{total}] {h}: Web {s}", "success" if ok else "error"))
                RESULTS_STORE['web'].append([es, sn, em, ts])
            elif task_type in ['web_enable', 'web_disable']:
                ok, msg = set_web_status(host, port, 1 if task_type=='web_enable' else 0)
                em = msg
                rc['ok' if ok else 'err'] += 1
                if not ok: es = "ERROR"
                self.root.after(0, lambda h=host, m=msg: self.log(f"[{i+1}/{total}] {h}: {m}", "success" if ok else "error"))
                RESULTS_STORE['web'].append([es, sn, em, ts])
            elif task_type == 'cloud_status':
                ok, st, _ = check_cloud_status(host, port)
                em = f"Облако {st}"
                rc['ok' if ok else 'err'] += 1
                if not ok: es = "ERROR"
                self.root.after(0, lambda h=host, s=st: self.log(f"[{i+1}/{total}] {h}: Облако {s}", "success" if ok else "error"))
                RESULTS_STORE['cloud'].append([es, sn, em, ts])
            elif task_type in ['cloud_enable', 'cloud_disable']:
                ok, msg = set_cloud_status(host, port, 1 if task_type=='cloud_enable' else 0)
                em = msg
                rc['ok' if ok else 'err'] += 1
                if not ok: es = "ERROR"
                self.root.after(0, lambda h=host, m=msg: self.log(f"[{i+1}/{total}] {h}: {m}", "success" if ok else "error"))
                RESULTS_STORE['cloud'].append([es, sn, em, ts])
            elif task_type == 'version':
                ok, msg, ver = check_version_single(host, port)
                ev = ver or "-"
                rc['ok' if ok else 'err'] += 1
                if not ok: es, em = "ERROR", msg
                self.root.after(0, lambda h=host, v=ver: self.log(f"[{i+1}/{total}] {h}: Версия: {v}", "success" if ok else "error"))
                RESULTS_STORE['version'].append([es, sn, ev, ts])
            elif task_type in ['archive_main', 'archive_subs', 'archive_both']:
                am, asub = "-", "-"
                lm = []
                if task_type in ['archive_main', 'archive_both']:
                    ok, _, d = check_archive_days(host, port, 'main')
                    am = str(round(float(d))) if d else "0"
                    lm.append(f"Main: {am} дн.")
                    rc['ok' if ok else 'err'] += 1
                if task_type in ['archive_subs', 'archive_both']:
                    ok, _, d = check_archive_days(host, port, 'subs')
                    asub = str(round(float(d))) if d else "0"
                    lm.append(f"Subs: {asub} дн.")
                    rc['ok' if ok else 'err'] += 1
                self.root.after(0, lambda m=', '.join(lm): self.log(f"[{i+1}/{total}] {sn}: {m}", "success"))
                if task_type == 'archive_main': asub = "-"
                if task_type == 'archive_subs': am = "-"
                RESULTS_STORE['archive'].append([es, sn, am, asub, ts])
            elif task_type == 'db_check':
                ok, msg, val = check_db_status(host, port)
                ev = val if val is not None else "-"
                rc['ok' if ok else 'err'] += 1
                if not ok: es, em = "ERROR", msg
                self.root.after(0, lambda h=host, m=msg: self.log(f"[{i+1}/{total}] {h}: {m}", "success" if ok else "error"))
                RESULTS_STORE['db'].append([es, sn, msg, ev, ts])
            elif task_type == 'pos_check':
                terms = get_pos_terminals_list(host, port)
                if terms is None:
                    rc['err'] += 1; es, em = "ERROR", "Нет соединения"
                    self.root.after(0, lambda h=host: self.log(f"[{i+1}/{total}] {h}: Ошибка получения списка POS", "error"))
                    RESULTS_STORE['pos'].append(["ERROR", sn, "-", "-", ts])
                elif not terms:
                    rc['ok'] += 1
                    self.root.after(0, lambda h=host: self.log(f"[{i+1}/{total}] {h}: POS терминалы не найдены", "warning"))
                    RESULTS_STORE['pos'].append(["OK", sn, "-", "-", ts])
                else:
                    sok, te = True, []
                    for tid in terms:
                        ok, td = check_pos_terminal(host, port, tid)
                        if ok:
                            n, p, pt, en = td.get('name','-'), td.get('port','-'), td.get('pos_type','-'), td.get('pos_enable')
                            es_ = "Вкл" if en==1 else ("Выкл" if en==0 else "?")
                            te.append(f"{tid} | {n}:{p}:{es_}")
                        else:
                            sok, te = False, te + [f"{tid} | ERR: {td.get('error','Ошибка')}"]
                    self.root.after(0, lambda m=f"[{i+1}/{total}] {host}: {' | '.join(te)}": self.log(m, "success" if sok else "error"))
                    RESULTS_STORE['pos'].append(["OK" if sok else "ERROR", sn, ", ".join(t.split(' | ')[0] for t in te), " | ".join(t.split(' | ')[1] if ' | ' in t else t for t in te), ts])
                    rc['ok' if sok else 'err'] += 1
            elif task_type == 'ip_check':
                ok, msg, ip = get_server_ip(host, port)
                ei = ip or "-"
                rc['ok' if ok else 'err'] += 1
                if not ok: es, em = "ERROR", msg
                self.root.after(0, lambda h=host, i=ei: self.log(f"[{i+1}/{total}] {h}: {i}", "success" if ok else "error"))
                RESULTS_STORE['ip'].append([es, sn, ei, ts])
            elif task_type == 'users_list':
                ud, _ = get_users_list(host, port)
                if ud is None:
                    rc['err'] += 1; es, em = "ERROR", "Нет соединения"
                    self.root.after(0, lambda h=host: self.log(f"[{i+1}/{total}] {h}: Ошибка получения списка", "error"))
                    RESULTS_STORE['users_list'].append(["ERROR", sn, "0", "Ошибка", ts])
                else:
                    rc['ok'] += 1
                    em = f"Найдено пользователей: {len(ud)}"
                    un = [f"{get_display_name(host, port, u)} ({u})" for u in ud]
                    us = ", ".join(un)
                    self.root.after(0, lambda h=host, c=len(ud): self.log(f"[{i+1}/{total}] {h}: {c} пользователей", "success"))
                    self.root.after(0, lambda u=us: self.log(f"  {u}", "user"))
                    RESULTS_STORE['users_list'].append(["OK", sn, str(len(ud)), us, ts])
            elif task_type == 'users_check':
                _, _, sv = check_version_single(host, port)
                ud, _ = get_users_list(host, port)
                if ud:
                    vi = f" (версия: {sv})" if sv else ""
                    self.root.after(0, lambda h=host, v=vi: self.log(f"\n[{i+1}/{total}] {h}{v}:", "header"))
                    for tgt in ['Admin', 'KRU', 'OPS', 'Manager']:
                        u = find_target_user(host, port, ud, tgt)
                        if not u:
                            self.root.after(0, lambda t=tgt: self.log(f"  ┌─ {t}", "user"))
                            self.root.after(0, lambda: self.log(f"  │  ❌ Не найден", "error"))
                            RESULTS_STORE['users_check'].append(["ERROR", sn, tgt, "-", "Не найден", ts])
                            continue
                        _, _, mm, _, _ = check_user_settings(host, port, u, tgt, sv)
                        self.root.after(0, lambda t=tgt, u=u: self.log(f"  ┌─ {t} ({u})", "user"))
                        if not mm:
                            self.root.after(0, lambda: self.log(f"  │  ✅ Все настройки соответствуют", "success"))
                            RESULTS_STORE['users_check'].append(["OK", sn, tgt, u, "Нет расхождений", ts])
                        else:
                            self.root.after(0, lambda m=len(mm): self.log(f"  │  ⚠️ Найдено расхождений: {m}", "warning"))
                            for x in mm: self.root.after(0, lambda x=x: self.log(f"  │    • {x}", "error"))
                            RESULTS_STORE['users_check'].append(["MISMATCH", sn, tgt, u, "; ".join(mm), ts])
                        rc['ok'] += 1
            self.root.after(0, lambda p=((i+1)/total)*100: self.progress_var.set(p))
        
        if not self.cancel_requested:
            fm = f"\n=== ИТОГИ ===\nУспешно: {rc['ok']}\nОшибок: {rc['err']}\n"
            self.root.after(0, lambda: self.log(fm, "header"))
            self.root.after(0, lambda: self.status_label.config(text="Задача завершена!"))
            self.root.after(0, lambda: messagebox.showinfo("Успех", "Все задачи выполнены успешно!") if rc['err']==0 else messagebox.showwarning("Завершено", f"Есть ошибки: {rc['err']}"))
        else:
            self.root.after(0, lambda: self.status_label.config(text="Задача отменена"))
        
        self.is_running, self.cancel_requested = False, False
        for b in [self.btn_ntp_check, self.btn_ntp_fix, self.btn_web_status, self.btn_web_enable, self.btn_web_disable, self.btn_cloud_status, self.btn_cloud_enable, self.btn_cloud_disable, self.btn_version, self.btn_archive_main, self.btn_archive_subs, self.btn_archive_both, self.btn_db_check, self.btn_pos_check, self.btn_ip_check, self.btn_users_list, self.btn_users_check, self.btn_copy_log, self.btn_save_excel, self.btn_clear_log]:
            try: self.root.after(0, lambda b=b: b.config(state='normal'))
            except: pass
        self.root.after(0, lambda: self.btn_cancel.config(state='disabled'))

    def get_servers(self):
        return [p for l in self.text_input.get("1.0", tk.END).splitlines() if (p:=parse_server_url(l))]

if __name__ == '__main__':
    root = tk.Tk()
    app = BKManagerApp(root)
    root.mainloop()